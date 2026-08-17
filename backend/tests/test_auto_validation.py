"""Auto-validation: rule matching engine, submission wiring for both DCB
and WRC, admin override + audit trail, and the export endpoint. See
app/services/auto_validation_service.py's module docstring for the
advisory-only design this suite is built around -- the central invariant
every test here protects is that auto-validation NEVER sets a bill/
incident's real `considered` decision."""

from datetime import date, datetime, timedelta, timezone
from io import BytesIO

from app.models.audit_log import AuditLog
from app.models.auto_validation import AutoValidationRule
from app.models.delayed_cash_billing import DelayedCashCaseResponse
from app.models.user import User
from app.models.weekly_revenue_closure import WeeklyRevenueCaseResponse
from app.services import auto_validation_service as service
from app.services import delayed_cash_penalty_service as dcb_svc
from app.services import weekly_revenue_closure_service as wrc_svc
from tests.conftest import TestingSessionLocal


def _register(client, username, email, password="password123"):
    return client.post("/auth/register", json={"username": username, "email": email, "password": password})


def _login(client, username, password="password123"):
    return client.post("/auth/login", json={"username": username, "password": password}).json()["access_token"]


def _auth(token):
    return {"Authorization": f"Bearer {token}"}


def _admin(client, username="av_admin", email="av_admin@example.com"):
    _register(client, username, email)
    db = TestingSessionLocal()
    try:
        db.query(User).filter(User.username == username).first().role = "Admin"
        db.commit()
    finally:
        db.close()
    return _login(client, username)


def _make_user(username):
    db = TestingSessionLocal()
    try:
        import bcrypt

        user = User(
            username=username, email=f"{username}@example.com",
            password_hash=bcrypt.hashpw(b"password123", bcrypt.gensalt()).decode(),
            role="Admin", is_active=True,
        )
        db.add(user)
        db.commit()
        db.refresh(user)
        return user
    finally:
        db.close()


def _evidence_file(name="proof.pdf", content=b"%PDF-1.4 fake evidence content"):
    return {"evidence": (name, BytesIO(content), "application/pdf")}


def _make_dcb_case(suffix: str):
    _make_user(f"av_dcb_setup{suffix}")
    db = TestingSessionLocal()
    try:
        user = db.query(User).filter(User.username == f"av_dcb_setup{suffix}").first()
        rule = dcb_svc.create_rule(db, rule_version=f"AV-DCB-{suffix}", created_by=user)
        dcb_svc.approve_rule(db, rule=rule, approver=user)
        batch = dcb_svc.create_upload_batch(
            db, period_start=date(2026, 7, 1), period_end=date(2026, 7, 31),
            source_filename=f"av-dcb-{suffix}.xlsx", rule=rule, uploaded_by=user,
        )
        bill_date = date(2026, 7, 1)
        created = bill_date + timedelta(days=2)
        raw = dcb_svc.RawBillInput(
            centre_code=f"AV-DCB-{suffix}", centre_name=f"AV DCB Center {suffix}",
            sales_bill=f"AV-DCB-{suffix}-1", bill_date=bill_date,
            bill_created_time=datetime.combine(created, datetime.min.time(), tzinfo=timezone.utc),
            created_date=created, source_day_difference=2,
        )
        dcb_svc.ingest_bills(db, batch=batch, rule=rule, raw_bills=[raw])
        center_penalties = dcb_svc.compute_center_penalties(db, batch=batch, rule=rule)
        return center_penalties[0].id
    finally:
        db.close()


def _make_wrc_case(suffix: str):
    _make_user(f"av_wrc_setup{suffix}")
    db = TestingSessionLocal()
    try:
        user = db.query(User).filter(User.username == f"av_wrc_setup{suffix}").first()
        rule = wrc_svc.create_rule(db, rule_version=f"AV-WRC-{suffix}", created_by=user)
        wrc_svc.approve_rule(db, rule=rule, approver=user)
        batch = wrc_svc.create_batch(
            db, period_start=date(2026, 7, 1), period_end=date(2026, 7, 7),
            week_label=f"AV Week {suffix}", rule=rule, created_by=user,
        )
        code = f"AV-WRC-{suffix}"
        raw = wrc_svc.RawBillIncidentInput(
            centre_code=code, centre_name=f"AV WRC Center {suffix}",
            incident_date=date(2026, 7, 2), mis_final_remark="bill_pending",
        )
        wrc_svc.record_bill_incidents(db, batch=batch, raw_incidents=[raw])
        return batch.id, code
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Pure matching engine
# ---------------------------------------------------------------------------


def test_evaluate_remark_matches_considered_rule():
    db = TestingSessionLocal()
    try:
        result = service.evaluate_remark(db, remark_text="Wrong bill created due to system error", engine="dcb")
        assert result.bucket == "considered"
        assert result.category == "Rebilling"
        assert result.matched_keyword == "Wrong bill created"
    finally:
        db.close()


def test_evaluate_remark_matches_not_considered_rule_and_carries_reason():
    db = TestingSessionLocal()
    try:
        result = service.evaluate_remark(db, remark_text="Sorry, staff on leave that day", engine="wrc")
        assert result.bucket == "not_considered"
        assert result.category == "Leave"
        assert result.reason == "Leaves & Holidays are not considered as per SOP"
    finally:
        db.close()


def test_evaluate_remark_with_no_rule_match_is_manual_check():
    db = TestingSessionLocal()
    try:
        result = service.evaluate_remark(db, remark_text="The patient census this week was unusual", engine="dcb")
        assert result.bucket == "manual_check"
    finally:
        db.close()


def test_evaluate_remark_matching_both_sides_is_manual_check():
    db = TestingSessionLocal()
    try:
        # "Proof will be shared later" (not_considered) + "Price modification" (considered).
        result = service.evaluate_remark(
            db, remark_text="Price modification case, proof will be shared later", engine="dcb"
        )
        assert result.bucket == "manual_check"
    finally:
        db.close()


def test_evaluate_remark_is_word_boundary_not_raw_substring():
    """"Holiday" must not fire on "holidaying" -- a raw substring match
    would incorrectly flag this as not_considered."""
    db = TestingSessionLocal()
    try:
        result = service.evaluate_remark(db, remark_text="We are not holidaying, the delay was a data glitch", engine="dcb")
        assert result.bucket == "manual_check"
    finally:
        db.close()


def test_evaluate_remark_empty_text_is_manual_check():
    db = TestingSessionLocal()
    try:
        result = service.evaluate_remark(db, remark_text="   ", engine="dcb")
        assert result.bucket == "manual_check"
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Rules CRUD + RBAC
# ---------------------------------------------------------------------------


def test_non_vigilance_role_cannot_list_or_create_rules(client):
    _admin(client)
    _register(client, "av_plain", "av_plain@example.com")
    db = TestingSessionLocal()
    try:
        db.query(User).filter(User.username == "av_plain").first().role = "Center Manager"
        db.commit()
    finally:
        db.close()
    token = _login(client, "av_plain")

    assert client.get("/auto-validation-rules", headers=_auth(token)).status_code == 403
    assert client.post(
        "/auto-validation-rules",
        json={"bucket": "considered", "category": "X", "keyword_phrase": "y", "decision_label": "Consider"},
        headers=_auth(token),
    ).status_code == 403


def test_admin_can_list_seeded_rules(client):
    admin_token = _admin(client, "av_admin2", "av_admin2@example.com")
    resp = client.get("/auto-validation-rules", headers=_auth(admin_token))
    assert resp.status_code == 200
    rules = resp.json()
    assert len(rules) == 49
    assert sum(1 for r in rules if r["bucket"] == "considered") == 16
    assert sum(1 for r in rules if r["bucket"] == "not_considered") == 33


def test_admin_can_create_and_deactivate_a_rule(client):
    admin_token = _admin(client, "av_admin3", "av_admin3@example.com")
    created = client.post(
        "/auto-validation-rules",
        json={
            "bucket": "not_considered", "category": "Test Category", "keyword_phrase": "totally made up phrase",
            "decision_label": "Not Considered", "reason": "Test reason",
        },
        headers=_auth(admin_token),
    )
    assert created.status_code == 201
    rule_id = created.json()["id"]
    assert created.json()["is_active"] is True

    deactivated = client.patch(
        f"/auto-validation-rules/{rule_id}/active", json={"is_active": False}, headers=_auth(admin_token)
    )
    assert deactivated.status_code == 200
    assert deactivated.json()["is_active"] is False

    db = TestingSessionLocal()
    try:
        rule = db.query(AutoValidationRule).filter(AutoValidationRule.id == rule_id).first()
        assert rule.is_active is False
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Submission wiring (public portal -> auto-validation, advisory only)
# ---------------------------------------------------------------------------


def test_dcb_submission_is_auto_validated_and_never_touches_real_decision(client):
    admin_token = _admin(client, "av_admin4", "av_admin4@example.com")
    cp_id = _make_dcb_case("4")
    link = client.post(f"/delayed-cash/center-penalties/{cp_id}/response-link", headers=_auth(admin_token)).json()

    resp = client.post(
        f"/public/delayed-cash/cases/{link['response_token']}/respond",
        data={
            "responder_name": "Test Manager", "responder_npid": "NP1", "responder_email": "m@example.com",
            "reason": "Sorry for delay, will not repeat",
        },
        files=_evidence_file(),
    )
    assert resp.status_code == 201
    response_id = resp.json()["id"]

    db = TestingSessionLocal()
    try:
        response = db.query(DelayedCashCaseResponse).filter(DelayedCashCaseResponse.id == response_id).first()
        assert response.auto_bucket == "not_considered"
        assert response.auto_category == "No Explanation"
        assert response.auto_reason == "No valid reason or No justification"
        assert response.auto_evaluated_at is not None
        # The real decision is untouched -- advisory only.
        bill = dcb_svc.list_bills_for_center_penalty(db, center_penalty=response.center_penalty)[0]
        assert bill.considered is None
    finally:
        db.close()

    listed = client.get("/delayed-cash/auto-validation", headers=_auth(admin_token)).json()
    assert any(r["id"] == response_id and r["effective_bucket"] == "not_considered" for r in listed)


def test_wrc_submission_is_auto_validated(client):
    admin_token = _admin(client, "av_admin5", "av_admin5@example.com")
    batch_id, code = _make_wrc_case("5")
    link = client.post(
        f"/weekly-revenue-closure/batches/{batch_id}/centers/{code}/response-link", headers=_auth(admin_token)
    ).json()

    resp = client.post(
        f"/public/weekly-revenue/cases/{link['response_token']}/respond",
        data={
            "responder_name": "M", "responder_npid": "NP1", "responder_email": "m@example.com",
            "reason": "IP bills pending for this patient",
        },
        files=_evidence_file(),
    )
    assert resp.status_code == 201
    response_id = resp.json()["id"]

    db = TestingSessionLocal()
    try:
        response = db.query(WeeklyRevenueCaseResponse).filter(WeeklyRevenueCaseResponse.id == response_id).first()
        assert response.auto_bucket == "considered"
        assert response.auto_category == "IP Bills Pending"
    finally:
        db.close()

    listed = client.get("/weekly-revenue-closure/auto-validation", headers=_auth(admin_token)).json()
    assert any(r["id"] == response_id and r["effective_bucket"] == "considered" for r in listed)


def test_reevaluate_single_response_updates_after_remark_unchanged_rules_change(client):
    """Simulates the on-demand re-run: deactivate the matching rule, then
    re-run -- the response should fall back to manual_check."""
    admin_token = _admin(client, "av_admin6", "av_admin6@example.com")
    cp_id = _make_dcb_case("6")
    link = client.post(f"/delayed-cash/center-penalties/{cp_id}/response-link", headers=_auth(admin_token)).json()
    resp = client.post(
        f"/public/delayed-cash/cases/{link['response_token']}/respond",
        data={"responder_name": "M", "responder_npid": "NP1", "responder_email": "m@example.com", "reason": "Working on it"},
        files=_evidence_file(),
    )
    response_id = resp.json()["id"]
    assert resp.json() is not None

    db = TestingSessionLocal()
    try:
        response = db.query(DelayedCashCaseResponse).filter(DelayedCashCaseResponse.id == response_id).first()
        assert response.auto_bucket == "not_considered"
        rule = (
            db.query(AutoValidationRule)
            .filter(AutoValidationRule.keyword_phrase == "Working on it")
            .first()
        )
        rule.is_active = False
        db.commit()
    finally:
        db.close()

    reevaluated = client.post(
        f"/delayed-cash/auto-validation/{response_id}/reevaluate", headers=_auth(admin_token)
    )
    assert reevaluated.status_code == 200
    assert reevaluated.json()["auto_bucket"] == "manual_check"

    # Restore the rule so this test doesn't leak state into others.
    db = TestingSessionLocal()
    try:
        rule = db.query(AutoValidationRule).filter(AutoValidationRule.keyword_phrase == "Working on it").first()
        rule.is_active = True
        db.commit()
    finally:
        db.close()


def test_reevaluate_all_skips_admin_overridden_responses(client):
    admin_token = _admin(client, "av_admin7", "av_admin7@example.com")
    cp_id = _make_dcb_case("7")
    link = client.post(f"/delayed-cash/center-penalties/{cp_id}/response-link", headers=_auth(admin_token)).json()
    resp = client.post(
        f"/public/delayed-cash/cases/{link['response_token']}/respond",
        data={"responder_name": "M", "responder_npid": "NP1", "responder_email": "m@example.com", "reason": "Team missed this one"},
        files=_evidence_file(),
    )
    response_id = resp.json()["id"]

    override = client.post(
        f"/delayed-cash/auto-validation/{response_id}/override",
        json={"bucket": "considered", "note": "Vigilance manually verified this was legitimate"},
        headers=_auth(admin_token),
    )
    assert override.status_code == 200
    assert override.json()["effective_bucket"] == "considered"
    assert override.json()["auto_bucket"] == "not_considered"  # original auto result preserved

    client.post("/delayed-cash/auto-validation/reevaluate-all", headers=_auth(admin_token))

    db = TestingSessionLocal()
    try:
        response = db.query(DelayedCashCaseResponse).filter(DelayedCashCaseResponse.id == response_id).first()
        # Still overridden -- reevaluate-all must never clobber a human decision.
        assert response.admin_override_bucket == "considered"
        assert response.auto_bucket == "not_considered"

        audit = (
            db.query(AuditLog)
            .filter(
                AuditLog.entity_type == "DelayedCashCaseResponse",
                AuditLog.entity_id == str(response_id),
                AuditLog.action == "auto_validation.overridden",
            )
            .first()
        )
        assert audit is not None
        assert audit.after_json["bucket"] == "considered"
    finally:
        db.close()


def test_override_invalid_bucket_rejected(client):
    admin_token = _admin(client, "av_admin8", "av_admin8@example.com")
    cp_id = _make_dcb_case("8")
    link = client.post(f"/delayed-cash/center-penalties/{cp_id}/response-link", headers=_auth(admin_token)).json()
    resp = client.post(
        f"/public/delayed-cash/cases/{link['response_token']}/respond",
        data={"responder_name": "M", "responder_npid": "NP1", "responder_email": "m@example.com", "reason": "no idea"},
        files=_evidence_file(),
    )
    response_id = resp.json()["id"]

    bad = client.post(
        f"/delayed-cash/auto-validation/{response_id}/override",
        json={"bucket": "not_a_real_bucket"},
        headers=_auth(admin_token),
    )
    assert bad.status_code == 422  # pydantic pattern validation rejects it before the service is even called


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------


def test_export_returns_a_workbook_with_expected_sheets(client):
    admin_token = _admin(client, "av_admin9", "av_admin9@example.com")
    cp_id = _make_dcb_case("9")
    link = client.post(f"/delayed-cash/center-penalties/{cp_id}/response-link", headers=_auth(admin_token)).json()
    client.post(
        f"/public/delayed-cash/cases/{link['response_token']}/respond",
        data={"responder_name": "M", "responder_npid": "NP1", "responder_email": "m@example.com", "reason": "Holiday, so delayed"},
        files=_evidence_file(),
    )

    resp = client.get("/auto-validation/export.xlsx", headers=_auth(admin_token))
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(resp.content))
    assert set(wb.sheetnames) == {"Raw Data", "By Center", "By Zone", "By Cluster"}
    assert wb["Raw Data"].max_row >= 2  # header + at least one data row
