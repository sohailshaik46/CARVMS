"""Tests for GET /delayed-cash/batches/{id}/export.xlsx -- the Data +
Penalty workbook regenerated from a batch's own computed rows."""

import io
from datetime import date, datetime
from decimal import Decimal

import openpyxl

from app.models.user import User
from tests.conftest import TestingSessionLocal

BILLS_DATA_HEADERS = [
    "CENTREID", "CENTRENAME", "SALESBILL", "BILLDATE", "bill_created_time",
    "created_date", "day_difference", "Center Remarks", "Penalty Remarks",
]


def _register(client, username, email, password="password123"):
    return client.post("/auth/register", json={"username": username, "email": email, "password": password})


def _login(client, username, password="password123"):
    return client.post("/auth/login", json={"username": username, "password": password}).json()["access_token"]


def _auth(token):
    return {"Authorization": f"Bearer {token}"}


def _admin(client, username, email):
    _register(client, username, email)
    db = TestingSessionLocal()
    try:
        db.query(User).filter(User.username == username).first().role = "Admin"
        db.commit()
    finally:
        db.close()
    return _login(client, username)


def _upload(client, token, rows, period_start="2026-07-01", period_end="2026-07-31"):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Bills Data"
    sheet.append(BILLS_DATA_HEADERS)
    for row in rows:
        sheet.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return client.post(
        "/delayed-cash/batches/upload",
        headers=_auth(token),
        data={"period_start": period_start, "period_end": period_end},
        files={"file": ("weekly.xlsx", io.BytesIO(buf.getvalue()), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


def test_export_workbook_round_trips_the_computed_penalty(client):
    admin_token = _admin(client, "dcb_export_admin", "dcb_export_admin@example.com")
    client.post("/delayed-cash/rules/activate-default", headers=_auth(admin_token))

    rows = [
        ["EXP-1", "Export Test Center", "BILL-1", date(2026, 7, 1), datetime(2026, 7, 2, 10, 0), date(2026, 7, 2), 1, None, None],
        ["EXP-1", "Export Test Center", "BILL-2", date(2026, 7, 1), datetime(2026, 7, 4, 10, 0), date(2026, 7, 4), 3, None, None],
    ]
    upload_resp = _upload(client, admin_token, rows)
    assert upload_resp.status_code == 201
    batch_id = upload_resp.json()["batch"]["id"]

    export = client.get(f"/delayed-cash/batches/{batch_id}/export.xlsx", headers=_auth(admin_token))
    assert export.status_code == 200
    assert export.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = openpyxl.load_workbook(io.BytesIO(export.content))
    assert set(wb.sheetnames) == {"Data", "Penalty"}

    data_rows = list(wb["Data"].iter_rows(values_only=True))
    assert data_rows[0][0] == "Center Code"
    data_codes = {r[0] for r in data_rows[1:]}
    assert data_codes == {"EXP-1"}
    assert len(data_rows) - 1 == 2  # both bills present

    penalty_rows = list(wb["Penalty"].iter_rows(values_only=True))
    assert penalty_rows[0][0] == "Center Code"
    center_row = next(r for r in penalty_rows[1:] if r[0] == "EXP-1")
    assert center_row[2] == 2  # total_bills
    assert Decimal(str(center_row[3])) == Decimal("400")  # (1 + 3) x 100
    grand_total_row = penalty_rows[-1]
    assert grand_total_row[0] == "Grand Total"
    assert Decimal(str(grand_total_row[3])) == Decimal("400")


def test_export_unknown_batch_404s(client):
    admin_token = _admin(client, "dcb_export_404", "dcb_export_404@example.com")
    resp = client.get("/delayed-cash/batches/999999/export.xlsx", headers=_auth(admin_token))
    assert resp.status_code == 404
