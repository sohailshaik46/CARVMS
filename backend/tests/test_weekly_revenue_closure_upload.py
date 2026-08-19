"""Tests for the Weekly Revenue Closure raw pending-list ingestion parser.

tests/fixtures/weekly_revenue_closure_week2_pending.json is extracted
programmatically (not hand-transcribed) from the real reference file
`July-26-Week2-closure pending List till 12-6-2026.xlsx`'s `Center wise`
sheet, plus that same file's own `Center Penalty` sheet as the expected
per-center incident-type counts -- both real, both from the same workbook,
so this is a genuine same-source reconciliation, not an invented example.
"""

import json
import os
from collections import defaultdict
from datetime import date, datetime

import openpyxl

from app.services import weekly_revenue_closure_upload_service as upload_service

FIXTURES_DIR = os.path.join(os.path.dirname(__file__), "fixtures")


def _load_fixture(name):
    with open(os.path.join(FIXTURES_DIR, name), "r", encoding="utf-8") as f:
        return json.load(f)


def _build_workbook(header, rows) -> bytes:
    import io

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Center wise"
    sheet.append(header)
    for row in rows:
        r = list(row)
        # date column (index 4) comes back as an ISO string from the JSON
        # fixture -- convert back to a real date so the parser exercises
        # the same code path it would against a real .xlsx upload.
        if r[4]:
            r[4] = date.fromisoformat(r[4])
        sheet.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# The real Week 2 fixture's own Date column spans exactly this range --
# using it as the batch period here means these reconciliation tests
# exercise the real out-of-period filter (it's just a no-op for this
# particular real file, since none of its rows fall outside it).
WEEK2_PERIOD_START = date(2026, 7, 1)
WEEK2_PERIOD_END = date(2026, 7, 12)


def test_parses_real_week2_pending_file_and_reconciles_exactly():
    fixture = _load_fixture("weekly_revenue_closure_week2_pending.json")
    content = _build_workbook(fixture["header"], fixture["rows"])

    incidents, excess_billed_count, out_of_period_count, skipped = upload_service.parse_pending_workbook(
        content, period_start=WEEK2_PERIOD_START, period_end=WEEK2_PERIOD_END,
    )

    assert skipped == []
    assert out_of_period_count == 0  # every row in this real file is genuinely within its own week
    assert excess_billed_count == fixture["expected_excess_billed_row_count"]

    actual_counts = defaultdict(lambda: defaultdict(int))
    for inc in incidents:
        actual_counts[inc.centre_code][inc.mis_final_remark] += 1

    expected = fixture["expected_center_incident_counts"]
    assert set(actual_counts.keys()) == set(expected.keys())
    for code, expected_types in expected.items():
        for type_key, expected_count in expected_types.items():
            assert actual_counts[code].get(type_key, 0) == expected_count, (code, type_key)


def test_parsed_incidents_have_no_remark_or_verdict_yet():
    """Ingestion produces pending incidents awaiting a center remark and a
    Vigilance verdict -- never a fabricated one."""
    fixture = _load_fixture("weekly_revenue_closure_week2_pending.json")
    content = _build_workbook(fixture["header"], fixture["rows"])

    incidents, _, _, _ = upload_service.parse_pending_workbook(
        content, period_start=WEEK2_PERIOD_START, period_end=WEEK2_PERIOD_END,
    )
    assert len(incidents) > 0
    for inc in incidents:
        assert inc.center_remarks is None
        assert inc.penalty_remarks is None


def test_excess_billed_rows_are_excluded_but_counted_not_silently_dropped():
    fixture = _load_fixture("weekly_revenue_closure_week2_pending.json")
    content = _build_workbook(fixture["header"], fixture["rows"])

    incidents, excess_billed_count, _, _ = upload_service.parse_pending_workbook(
        content, period_start=WEEK2_PERIOD_START, period_end=WEEK2_PERIOD_END,
    )
    assert excess_billed_count == 16
    # None of the parsed (penalty-eligible) incidents are the excess type.
    assert all(inc.mis_final_remark != upload_service.EXCESS_BILLED_TYPE for inc in incidents)


def test_missing_required_column_raises_clear_error():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Center wise"
    sheet.append(["Something", "Else", "Entirely"])
    import io

    buf = io.BytesIO()
    wb.save(buf)

    import pytest

    with pytest.raises(ValueError, match="missing required column"):
        upload_service.parse_pending_workbook(buf.getvalue(), period_start=date(2026, 7, 1), period_end=date(2026, 7, 12))


def test_unparseable_date_is_skipped_not_fatal():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Center wise"
    sheet.append(["Zone", "Cluster", "Center Code", "Center Name", "Date", "Billed Sessions", "Daily Report", "Variance", "Remark", "Final Remarks"])
    sheet.append(["South", "Test Cluster", "TEST-C", "Test Center", "not-a-date", 5, 4, -1, "1 Bill pending", "Bill Pending"])
    sheet.append(["South", "Test Cluster", "TEST-C", "Test Center", date(2026, 7, 5), 5, 4, -1, "1 Bill pending", "Bill Pending"])
    import io

    buf = io.BytesIO()
    wb.save(buf)

    incidents, _, _, skipped = upload_service.parse_pending_workbook(
        buf.getvalue(), period_start=date(2026, 7, 1), period_end=date(2026, 7, 12),
    )
    assert len(incidents) == 1
    assert len(skipped) == 1
    assert "date" in skipped[0].reason.lower()


def test_rows_from_a_prior_week_are_excluded_and_counted_not_ingested():
    """The exact scenario reported: an exported pending list still carries
    the previous week's rows too -- those must not become incidents in
    THIS batch (already ingested/penalized by the earlier week's own
    upload)."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Center wise"
    sheet.append(["Zone", "Cluster", "Center Code", "Center Name", "Date", "Billed Sessions", "Daily Report", "Variance", "Remark", "Final Remarks"])
    # Week 2 Aug'26 = 10-16 Aug'26. Two rows genuinely belong to this week;
    # two are leftover from the prior week (3-9 Aug'26) still present in
    # the same export.
    sheet.append(["South", "Test Cluster", "TEST-C", "Test Center", date(2026, 8, 10), 5, 4, -1, "1 Bill pending", "Bill Pending"])
    sheet.append(["South", "Test Cluster", "TEST-C", "Test Center", date(2026, 8, 16), 5, 4, -1, "1 Bill pending", "Bill Pending"])
    sheet.append(["South", "Test Cluster", "TEST-C", "Test Center", date(2026, 8, 9), 5, 4, -1, "1 Bill pending", "Bill Pending"])
    sheet.append(["South", "Test Cluster", "TEST-C", "Test Center", date(2026, 8, 3), 5, 4, -1, "1 Bill pending", "Bill Pending"])
    import io

    buf = io.BytesIO()
    wb.save(buf)

    incidents, excess_billed_count, out_of_period_count, skipped = upload_service.parse_pending_workbook(
        buf.getvalue(), period_start=date(2026, 8, 10), period_end=date(2026, 8, 16),
    )
    assert len(incidents) == 2
    assert {inc.incident_date for inc in incidents} == {date(2026, 8, 10), date(2026, 8, 16)}
    assert out_of_period_count == 2
    assert excess_billed_count == 0
    assert skipped == []  # not an error -- excluded via the dedicated counter, same as excess-billed


def test_out_of_period_rows_are_not_reported_as_skipped_errors():
    """Mirrors the excess-billed convention exactly -- an out-of-period row
    is a normal, expected exclusion, not something that should look like
    a data problem in the skipped-rows report."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Center wise"
    sheet.append(["Zone", "Cluster", "Center Code", "Center Name", "Date", "Billed Sessions", "Daily Report", "Variance", "Remark", "Final Remarks"])
    sheet.append(["South", "Test Cluster", "TEST-C", "Test Center", date(2026, 8, 1), 5, 4, -1, "1 Bill pending", "Bill Pending"])
    import io

    buf = io.BytesIO()
    wb.save(buf)

    incidents, _, out_of_period_count, skipped = upload_service.parse_pending_workbook(
        buf.getvalue(), period_start=date(2026, 8, 10), period_end=date(2026, 8, 16),
    )
    assert incidents == []
    assert out_of_period_count == 1
    assert skipped == []
