"""Tests for the Weekly Revenue Closure API: upload, review queue,
mark-no-remark-received, close batch, and the batch summary KPI endpoint.

Builds a real in-memory .xlsx matching the proven `Center wise` pending-
list format (see weekly_revenue_closure_upload_service.py) rather than
calling the calculator service directly -- this exercises the actual
upload/parsing layer end to end, same convention as
test_delayed_cash_upload.py.
"""

import io
from datetime import date

import openpyxl

from app.models.user import User
from app.services import weekly_revenue_closure_service as calc_svc
from tests.conftest import TestingSessionLocal

PENDING_HEADERS = [
    "Zone", "Cluster", "Center Code", "Center Name", "Date",
    "Billed Sessions", "Daily Report", "Variance", "Remark", "Final Remarks",
]


def _build_pending_workbook(rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Center wise"
    sheet.append(PENDING_HEADERS)
    for row in rows:
        sheet.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _register(client, username, email, password="password123"):
    return client.post("/auth/register", json={"username": username, "email": email, "password": password})


def _login(client, username, password="password123"):
    return client.post("/auth/login", json={"username": username, "password": password}).json()["access_token"]


def _auth(token):
    return {"Authorization": f"Bearer {token}"}


def _admin(client, username, email):
    _register(client, username, email)
    db = TestingSessionLocal()
    try:
        db.query(User).filter(User.username == username).first().role = "Admin"
        db.commit()
    finally:
        db.close()
    return _login(client, username)


def _ensure_approved_rule(rule_version: str):
    db = TestingSessionLocal()
    try:
        user = db.query(User).filter(User.username == "wrc_api_rule_setup").first()
        if user is None:
            import bcrypt

            user = User(
                username="wrc_api_rule_setup",
                email="wrc_api_rule_setup@example.com",
                password_hash=bcrypt.hashpw(b"password123", bcrypt.gensalt()).decode(),
                role="Admin",
                is_active=True,
            )
            db.add(user)
            db.commit()
            db.refresh(user)
        rule = calc_svc.create_rule(db, rule_version=rule_version, created_by=user)
        calc_svc.approve_rule(db, rule=rule, approver=user)
    finally:
        db.close()


def _upload(client, token, rows, period_start="2026-07-01", period_end="2026-07-07", week_label="Week 99 - Test"):
    content = _build_pending_workbook(rows)
    return client.post(
        "/weekly-revenue-closure/batches/upload",
        headers=_auth(token),
        data={"period_start": period_start, "period_end": period_end, "week_label": week_label},
        files={"file": ("pending.xlsx", io.BytesIO(content), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


# ---------------------------------------------------------------------------
# RBAC + upload
# ---------------------------------------------------------------------------


def test_upload_requires_vigilance_role(client):
    _ensure_approved_rule("WRC-API-RBAC")
    _register(client, "wrc_plain", "wrc_plain@example.com")
    db = TestingSessionLocal()
    try:
        db.query(User).filter(User.username == "wrc_plain").first().role = "Center Manager"
        db.commit()
    finally:
        db.close()
    token = _login(client, "wrc_plain")

    resp = _upload(client, token, [["South", "Test Cluster", "TEST-C", "Test Center", date(2026, 7, 5), 5, 4, -1, "1 Bill pending", "Bill Pending"]])
    assert resp.status_code == 403


def test_upload_without_approved_rule_returns_400(client, monkeypatch):
    admin_token = _admin(client, "wrc_admin_norule", "wrc_admin_norule@example.com")

    def _raise(*args, **kwargs):
        raise calc_svc.NoApprovedRuleError("No approved rule -- test override.")

    monkeypatch.setattr("app.api.weekly_revenue_closure.calc_service.get_active_rule", _raise)

    resp = _upload(client, admin_token, [["South", "Test Cluster", "TEST-C", "Test Center", date(2026, 7, 5), 5, 4, -1, "1 Bill pending", "Bill Pending"]])
    assert resp.status_code == 400
    assert "no approved" in resp.json()["detail"].lower()


def test_upload_ingests_incidents_and_excludes_excess_billed(client):
    _ensure_approved_rule("WRC-API-UPLOAD")
    admin_token = _admin(client, "wrc_admin_upload", "wrc_admin_upload@example.com")

    rows = [
        ["South", "Test Cluster", "WRC-API-1", "API Test Center 1", date(2026, 7, 5), 5, 4, -1, "1 Bill pending", "Bill Pending"],
        ["South", "Test Cluster", "WRC-API-1", "API Test Center 1", date(2026, 7, 6), 42, 41, 1, "Excess Billing", "Excess Billing  or  Incorrect Daily report"],
        ["West", "Other Cluster", "WRC-API-2", "API Test Center 2", date(2026, 7, 5), 3, None, 3, "Daily Report was not sent", "Daily Report not sent"],
    ]
    resp = _upload(client, admin_token, rows)
    assert resp.status_code == 201
    body = resp.json()
    assert body["incidents_ingested"] == 2
    assert body["excess_billed_row_count"] == 1
    assert body["skipped_rows"] == []
    assert body["batch"]["week_label"] == "Week 99 - Test"
    assert body["batch"]["status"] == "open"


# ---------------------------------------------------------------------------
# Review queue
# ---------------------------------------------------------------------------


def test_review_queue_and_review_decision_flow(client):
    _ensure_approved_rule("WRC-API-REVIEW")
    admin_token = _admin(client, "wrc_admin_review", "wrc_admin_review@example.com")

    rows = [
        ["South", "Review Cluster", "WRC-REV-1", "Review Test Center", date(2026, 7, 5), 5, 4, -1, "1 Bill pending", "Bill Pending"],
    ]
    upload_resp = _upload(client, admin_token, rows, week_label="Week Review")
    batch_id = upload_resp.json()["batch"]["id"]

    queue = client.get("/weekly-revenue-closure/bills/review-queue", params={"batch_id": batch_id}, headers=_auth(admin_token))
    assert queue.status_code == 200
    assert len(queue.json()) == 1
    incident_id = queue.json()[0]["id"]
    assert queue.json()[0]["considered"] is None

    review = client.post(
        f"/weekly-revenue-closure/bills/{incident_id}/review",
        json={"decision": "not_considered", "center_remarks": "Center never provided proof."},
        headers=_auth(admin_token),
    )
    assert review.status_code == 200
    assert review.json()["considered"] == "not_considered"
    assert review.json()["center_remarks"] == "Center never provided proof."

    queue_after = client.get("/weekly-revenue-closure/bills/review-queue", params={"batch_id": batch_id}, headers=_auth(admin_token))
    assert queue_after.json() == []


def test_review_invalid_decision_rejected(client):
    _ensure_approved_rule("WRC-API-INVALID")
    admin_token = _admin(client, "wrc_admin_invalid", "wrc_admin_invalid@example.com")
    rows = [["South", "C", "WRC-INV-1", "Invalid Test Center", date(2026, 7, 5), 5, 4, -1, "1 Bill pending", "Bill Pending"]]
    upload_resp = _upload(client, admin_token, rows)
    incident_id = client.get(
        "/weekly-revenue-closure/bills/review-queue",
        params={"batch_id": upload_resp.json()["batch"]["id"]},
        headers=_auth(admin_token),
    ).json()[0]["id"]

    resp = client.post(
        f"/weekly-revenue-closure/bills/{incident_id}/review", json={"decision": "maybe"}, headers=_auth(admin_token)
    )
    assert resp.status_code == 400


def test_review_unknown_incident_404s(client):
    admin_token = _admin(client, "wrc_admin_404", "wrc_admin_404@example.com")
    resp = client.post(
        "/weekly-revenue-closure/bills/999999/review", json={"decision": "considered"}, headers=_auth(admin_token)
    )
    assert resp.status_code == 404


# ---------------------------------------------------------------------------
# Mark no-remark-received + close batch + summary
# ---------------------------------------------------------------------------


def test_mark_no_remark_received_and_close_batch_end_to_end(client):
    _ensure_approved_rule("WRC-API-CLOSE")
    admin_token = _admin(client, "wrc_admin_close", "wrc_admin_close@example.com")

    rows = [
        # Reviewed and rejected -> feeds not_considered_penalty + cluster escalation.
        ["South", "Close Cluster", "WRC-CLOSE-1", "Close Test Center 1", date(2026, 7, 5), 5, 4, -1, "1 Bill pending", "Bill Pending"],
        # Never responded -> gets marked no-remark-received.
        ["South", "Close Cluster", "WRC-CLOSE-2", "Close Test Center 2", date(2026, 7, 6), 3, None, 3, "Daily Report was not sent", "Daily Report not sent"],
    ]
    upload_resp = _upload(client, admin_token, rows, week_label="Week Close")
    batch_id = upload_resp.json()["batch"]["id"]
    queue = client.get(
        "/weekly-revenue-closure/bills/review-queue", params={"batch_id": batch_id}, headers=_auth(admin_token)
    ).json()
    by_code = {b["centre_code"]: b for b in queue}

    client.post(
        f"/weekly-revenue-closure/bills/{by_code['WRC-CLOSE-1']['id']}/review",
        json={"decision": "not_considered"},
        headers=_auth(admin_token),
    )
    no_remark = client.post(
        f"/weekly-revenue-closure/bills/{by_code['WRC-CLOSE-2']['id']}/mark-no-remark-received",
        headers=_auth(admin_token),
    )
    assert no_remark.status_code == 200
    assert no_remark.json()["centre_code"] == "WRC-CLOSE-2"
    assert no_remark.json()["incident_count"] == 1

    close = client.post(f"/weekly-revenue-closure/batches/{batch_id}/close", headers=_auth(admin_token))
    assert close.status_code == 200
    close_body = close.json()
    assert close_body["batch"]["status"] == "closed"
    center_penalties = {cp["centre_code"]: cp for cp in close_body["center_penalties"]}
    assert float(center_penalties["WRC-CLOSE-1"]["not_considered_penalty"]) == 0.0625
    assert float(center_penalties["WRC-CLOSE-2"]["no_remark_penalty"]) == 0.0625

    summary = client.get(f"/weekly-revenue-closure/batches/{batch_id}/summary", headers=_auth(admin_token))
    assert summary.status_code == 200
    summary_body = summary.json()
    assert summary_body["total_incidents"] == 2
    assert summary_body["not_considered_count"] == 1
    assert summary_body["no_remark_center_count"] == 1
    assert summary_body["centers_penalized"] == 2
    assert float(summary_body["total_center_penalty_rate"]) == 0.125


def test_export_workbook_round_trips_the_computed_penalties(client):
    _ensure_approved_rule("WRC-API-EXPORT")
    admin_token = _admin(client, "wrc_admin_export", "wrc_admin_export@example.com")

    rows = [
        ["South", "Export Cluster", "WRC-EXP-1", "Export Test Center 1", date(2026, 7, 5), 5, 4, -1, "1 Bill pending", "Bill Pending"],
        ["West", "Export Cluster 2", "WRC-EXP-2", "Export Test Center 2", date(2026, 7, 6), 3, None, 3, "Daily Report was not sent", "Daily Report not sent"],
    ]
    upload_resp = _upload(client, admin_token, rows, week_label="Week Export")
    batch_id = upload_resp.json()["batch"]["id"]
    queue = client.get(
        "/weekly-revenue-closure/bills/review-queue", params={"batch_id": batch_id}, headers=_auth(admin_token)
    ).json()
    by_code = {b["centre_code"]: b for b in queue}

    client.post(
        f"/weekly-revenue-closure/bills/{by_code['WRC-EXP-1']['id']}/review",
        json={"decision": "not_considered"},
        headers=_auth(admin_token),
    )
    client.post(
        f"/weekly-revenue-closure/bills/{by_code['WRC-EXP-2']['id']}/mark-no-remark-received",
        headers=_auth(admin_token),
    )
    client.post(f"/weekly-revenue-closure/batches/{batch_id}/close", headers=_auth(admin_token))

    export = client.get(f"/weekly-revenue-closure/batches/{batch_id}/export.xlsx", headers=_auth(admin_token))
    assert export.status_code == 200
    assert export.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(export.content))
    assert set(wb.sheetnames) == {"Sheet1", "Penalties", "Data"}

    sheet1_rows = list(wb["Sheet1"].iter_rows(values_only=True))
    sheet1_codes = {
        r[0] for r in sheet1_rows if r[0] not in (None, "Count of Center Code", "Center Code", "Grand Total")
    }
    assert sheet1_codes == {"WRC-EXP-1", "WRC-EXP-2"}

    penalties_text = [cell for row in wb["Penalties"].iter_rows(values_only=True) for cell in row]
    assert "WRC-EXP-1" in penalties_text  # section 1 (not considered)
    assert "WRC-EXP-2" in penalties_text  # section 2 (no remark)
    assert "Export Cluster" in penalties_text  # cluster rollup name present

    data_rows = list(wb["Data"].iter_rows(values_only=True))
    data_codes = {r[3] for r in data_rows[1:]}
    assert data_codes == {"WRC-EXP-1"}  # Data sheet only has remark-received incidents, per the proven format


def test_export_unknown_batch_404s(client):
    admin_token = _admin(client, "wrc_admin_export404", "wrc_admin_export404@example.com")
    resp = client.get("/weekly-revenue-closure/batches/999999/export.xlsx", headers=_auth(admin_token))
    assert resp.status_code == 404


def test_list_batches_and_get_batch(client):
    _ensure_approved_rule("WRC-API-LIST")
    admin_token = _admin(client, "wrc_admin_list", "wrc_admin_list@example.com")
    rows = [["South", "C", "WRC-LIST-1", "List Test Center", date(2026, 7, 5), 5, 4, -1, "1 Bill pending", "Bill Pending"]]
    upload_resp = _upload(client, admin_token, rows, week_label="Week List")
    batch_id = upload_resp.json()["batch"]["id"]

    listing = client.get("/weekly-revenue-closure/batches", headers=_auth(admin_token))
    assert listing.status_code == 200
    assert any(b["id"] == batch_id for b in listing.json())

    detail = client.get(f"/weekly-revenue-closure/batches/{batch_id}", headers=_auth(admin_token))
    assert detail.status_code == 200
    assert detail.json()["week_label"] == "Week List"

    assert client.get("/weekly-revenue-closure/batches/999999", headers=_auth(admin_token)).status_code == 404
    assert client.get("/weekly-revenue-closure/batches").status_code == 401
