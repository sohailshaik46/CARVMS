"""Tests for the Delayed Cash Billing upload + publishing pipeline --
POST /delayed-cash/batches/upload, GET /delayed-cash/batches[/{id}], and
POST /delayed-cash/batches/{id}/publish.

Builds real .xlsx workbooks in-memory (via openpyxl) matching the proven
'Bills Data' sheet layout from
docs/CARVMS_DELAYED_CASH_PENALTY_FORMULA_ANALYSIS.md, rather than calling
the calculator service directly -- this suite exercises the actual parsing
layer, not just the already-tested calculator underneath it.
"""

import io
from datetime import date, datetime, timedelta
from decimal import Decimal

import openpyxl

from app.models.user import User
from app.services import delayed_cash_penalty_service as calc_svc
from tests.conftest import TestingSessionLocal

BILLS_DATA_HEADERS = [
    "CENTREID",
    "CENTRENAME",
    "SALESBILL",
    "BILLDATE",
    "bill_created_time",
    "created_date",
    "day_difference",
    "Center Remarks",
    "Penalty Remarks",
]


def _build_workbook(rows: list[list]) -> bytes:
    """rows are raw values in BILLS_DATA_HEADERS order (any row may omit
    trailing columns)."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Bills Data"
    sheet.append(BILLS_DATA_HEADERS)
    for row in rows:
        sheet.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _register(client, username, email, password="password123"):
    return client.post("/auth/register", json={"username": username, "email": email, "password": password})


def _login(client, username, password="password123"):
    return client.post("/auth/login", json={"username": username, "password": password}).json()["access_token"]


def _auth(token):
    return {"Authorization": f"Bearer {token}"}


def _admin(client, username, email):
    _register(client, username, email)
    db = TestingSessionLocal()
    try:
        db.query(User).filter(User.username == username).first().role = "Admin"
        db.commit()
    finally:
        db.close()
    return _login(client, username)


def _ensure_approved_rule(rule_version: str):
    db = TestingSessionLocal()
    try:
        user = db.query(User).filter(User.username == "dcb_upload_rule_setup").first()
        if user is None:
            import bcrypt

            user = User(
                username="dcb_upload_rule_setup",
                email="dcb_upload_rule_setup@example.com",
                password_hash=bcrypt.hashpw(b"password123", bcrypt.gensalt()).decode(),
                role="Admin",
                is_active=True,
            )
            db.add(user)
            db.commit()
            db.refresh(user)
        rule = calc_svc.create_rule(db, rule_version=rule_version, created_by=user)
        calc_svc.approve_rule(db, rule=rule, approver=user)
    finally:
        db.close()


def _upload(client, token, rows, filename="weekly.xlsx", period_start="2026-07-01", period_end="2026-07-31"):
    content = _build_workbook(rows)
    return client.post(
        "/delayed-cash/batches/upload",
        headers=_auth(token),
        data={"period_start": period_start, "period_end": period_end},
        files={"file": (filename, io.BytesIO(content), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


# ---------------------------------------------------------------------------
# RBAC
# ---------------------------------------------------------------------------


def test_non_vigilance_role_cannot_upload(client):
    _ensure_approved_rule("DCB-UPLOAD-RBAC")
    _register(client, "dcbu_plain", "dcbu_plain@example.com")
    db = TestingSessionLocal()
    try:
        db.query(User).filter(User.username == "dcbu_plain").first().role = "Center Manager"
        db.commit()
    finally:
        db.close()
    token = _login(client, "dcbu_plain")

    resp = _upload(client, token, [["1-C", "Test Center", "S-1", "2026-07-01", "2026-07-02", "2026-07-02", 1]])
    assert resp.status_code == 403


def test_upload_requires_auth(client):
    content = _build_workbook([])
    resp = client.post(
        "/delayed-cash/batches/upload",
        data={"period_start": "2026-07-01", "period_end": "2026-07-31"},
        files={"file": ("weekly.xlsx", io.BytesIO(content), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 401


# ---------------------------------------------------------------------------
# No approved rule yet
# ---------------------------------------------------------------------------


def test_upload_without_approved_rule_returns_clear_400(client, monkeypatch):
    """Other test files in this shared test DB may have already approved a
    rule by the time this one runs, so the endpoint's own guard is exercised
    deterministically via monkeypatch instead of relying on true DB
    emptiness (mirrors the pattern in test_org_sheet_sync.py)."""
    admin_token = _admin(client, "dcbu_admin_norule", "dcbu_admin_norule@example.com")

    from app.services import delayed_cash_penalty_service as calc_svc_module

    def _raise_no_rule(*args, **kwargs):
        raise calc_svc_module.NoApprovedRuleError("No approved DelayedCashPenaltyRule exists yet -- test override.")

    monkeypatch.setattr("app.api.delayed_cash.calc_service.get_active_rule", _raise_no_rule)

    resp = _upload(client, admin_token, [["1-C", "Test Center", "S-1", date(2026, 7, 1), datetime(2026, 7, 2, 9, 0), date(2026, 7, 2), 1]])
    assert resp.status_code == 400
    assert "no approved" in resp.json()["detail"].lower()


# ---------------------------------------------------------------------------
# Successful ingestion
# ---------------------------------------------------------------------------


def test_upload_ingests_bills_and_computes_center_penalties(client):
    _ensure_approved_rule("DCB-UPLOAD-OK")
    admin_token = _admin(client, "dcbu_admin_ok", "dcbu_admin_ok@example.com")

    rows = [
        ["900-X-C", "Upload Test Center", "UP-1", date(2026, 7, 1), datetime(2026, 7, 2, 10, 0), date(2026, 7, 2), 1, None, None],
        ["900-X-C", "Upload Test Center", "UP-2", date(2026, 7, 1), datetime(2026, 7, 4, 10, 0), date(2026, 7, 4), 3, None, None],
    ]
    resp = _upload(client, admin_token, rows)
    assert resp.status_code == 201
    body = resp.json()

    assert body["batch"]["status"] == "uploaded"
    assert body["batch"]["source_filename"] == "weekly.xlsx"
    assert body["skipped_rows"] == []

    centers = {cp["centre_code"]: cp for cp in body["center_penalties"]}
    assert "900-X-C" in centers
    cp = centers["900-X-C"]
    assert cp["total_bills"] == 2
    assert Decimal(cp["calculated_penalty"]) == Decimal("400")  # (1 + 3) x 100
    assert cp["penalty_status"] == "published"


def test_upload_finds_data_on_a_sheet_other_than_bills_data(client):
    """Real source files aren't consistent about which sheet the data
    lands on -- proves the parser searches every sheet rather than
    assuming a fixed name/position."""
    _ensure_approved_rule("DCB-UPLOAD-ANYSHEET")
    admin_token = _admin(client, "dcbu_admin_anysheet", "dcbu_admin_anysheet@example.com")

    wb = openpyxl.Workbook()
    decoy = wb.active
    decoy.title = "Summary"
    decoy.append(["Not", "The", "Data"])

    data_sheet = wb.create_sheet("Sheet2")
    data_sheet.append(BILLS_DATA_HEADERS)
    data_sheet.append(
        ["902-Z-C", "Any Sheet Center", "AS-1", date(2026, 7, 1), datetime(2026, 7, 3, 9, 0), date(2026, 7, 3), 2, None, None]
    )
    buf = io.BytesIO()
    wb.save(buf)

    resp = client.post(
        "/delayed-cash/batches/upload",
        headers=_auth(admin_token),
        data={"period_start": "2026-07-01", "period_end": "2026-07-31"},
        files={"file": ("anysheet.xlsx", io.BytesIO(buf.getvalue()), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 201
    body = resp.json()
    assert body["skipped_rows"] == []
    centers = {cp["centre_code"]: cp for cp in body["center_penalties"]}
    assert "902-Z-C" in centers
    assert Decimal(centers["902-Z-C"]["calculated_penalty"]) == Decimal("200")  # 2 x 100


def test_upload_skips_bad_rows_without_aborting_the_batch(client):
    _ensure_approved_rule("DCB-UPLOAD-SKIP")
    admin_token = _admin(client, "dcbu_admin_skip", "dcbu_admin_skip@example.com")

    rows = [
        # valid
        ["901-Y-C", "Skip Test Center", "SK-1", date(2026, 7, 1), datetime(2026, 7, 2, 9, 0), date(2026, 7, 2), 1, None, None],
        # missing centre code -> skipped
        ["", "No Code Center", "SK-2", date(2026, 7, 1), datetime(2026, 7, 3, 9, 0), date(2026, 7, 3), 2, None, None],
        # duplicate sales bill (same as row 1) -> skipped
        ["901-Y-C", "Skip Test Center", "SK-1", date(2026, 7, 1), datetime(2026, 7, 5, 9, 0), date(2026, 7, 5), 4, None, None],
        # non-numeric day_difference -> skipped
        ["901-Y-C", "Skip Test Center", "SK-3", date(2026, 7, 1), datetime(2026, 7, 6, 9, 0), date(2026, 7, 6), "N/A", None, None],
        # unparseable created_date -> skipped
        ["901-Y-C", "Skip Test Center", "SK-4", date(2026, 7, 1), "not-a-date", "also-not-a-date", 5, None, None],
    ]
    resp = _upload(client, admin_token, rows)
    assert resp.status_code == 201
    body = resp.json()

    assert len(body["skipped_rows"]) == 4
    reasons = " | ".join(r["reason"] for r in body["skipped_rows"])
    assert "Missing Center Code" in reasons
    assert "Duplicate Sales Bill" in reasons
    assert "Non-numeric day_difference" in reasons
    assert "Unparseable" in reasons

    centers = {cp["centre_code"]: cp for cp in body["center_penalties"]}
    cp = centers["901-Y-C"]
    # Only the one genuinely valid row survived.
    assert cp["total_bills"] == 1
    assert Decimal(cp["calculated_penalty"]) == Decimal("100")


def test_rows_from_a_prior_week_are_excluded_and_counted_not_ingested(client):
    """The real bug this guards against: the source workbook for week N
    routinely carries a few rows from week N-1 (already ingested and
    penalized in that prior week's own upload) -- BILLDATE outside this
    batch's own period_start/period_end must be excluded, not
    double-counted into this batch too."""
    _ensure_approved_rule("DCB-UPLOAD-OUTOFPERIOD")
    admin_token = _admin(client, "dcbu_admin_oop", "dcbu_admin_oop@example.com")

    rows = [
        # inside the period (2026-07-01 .. 2026-07-31)
        ["909-G-C", "Out Of Period Center", "OOP-1", date(2026, 7, 10), datetime(2026, 7, 12, 9, 0), date(2026, 7, 12), 2, None, None],
        ["909-G-C", "Out Of Period Center", "OOP-2", date(2026, 7, 20), datetime(2026, 7, 22, 9, 0), date(2026, 7, 22), 2, None, None],
        # from the PRIOR period -- must be excluded
        ["909-G-C", "Out Of Period Center", "OOP-3", date(2026, 6, 25), datetime(2026, 6, 27, 9, 0), date(2026, 6, 27), 2, None, None],
        ["909-G-C", "Out Of Period Center", "OOP-4", date(2026, 6, 28), datetime(2026, 6, 30, 9, 0), date(2026, 6, 30), 2, None, None],
    ]
    resp = _upload(client, admin_token, rows, period_start="2026-07-01", period_end="2026-07-31")
    assert resp.status_code == 201
    body = resp.json()

    assert body["out_of_period_row_count"] == 2
    assert body["skipped_rows"] == []  # an expected exclusion, not a data error

    centers = {cp["centre_code"]: cp for cp in body["center_penalties"]}
    cp = centers["909-G-C"]
    assert cp["total_bills"] == 2  # only the two in-period rows


def test_upload_missing_required_column_returns_400(client):
    _ensure_approved_rule("DCB-UPLOAD-BADHEADER")
    admin_token = _admin(client, "dcbu_admin_badheader", "dcbu_admin_badheader@example.com")

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Bills Data"
    sheet.append(["Something", "Else", "Entirely"])
    buf = io.BytesIO()
    wb.save(buf)

    resp = client.post(
        "/delayed-cash/batches/upload",
        headers=_auth(admin_token),
        data={"period_start": "2026-07-01", "period_end": "2026-07-31"},
        files={"file": ("bad.xlsx", io.BytesIO(buf.getvalue()), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 400
    assert "no sheet in this workbook has every required column" in resp.json()["detail"].lower()


# ---------------------------------------------------------------------------
# Batch listing
# ---------------------------------------------------------------------------


def test_list_and_get_batch(client):
    _ensure_approved_rule("DCB-UPLOAD-LIST")
    admin_token = _admin(client, "dcbu_admin_list", "dcbu_admin_list@example.com")

    rows = [["902-Z-C", "List Test Center", "LT-1", date(2026, 7, 1), datetime(2026, 7, 2, 9, 0), date(2026, 7, 2), 1, None, None]]
    upload_resp = _upload(client, admin_token, rows)
    batch_id = upload_resp.json()["batch"]["id"]

    listing = client.get("/delayed-cash/batches", headers=_auth(admin_token))
    assert listing.status_code == 200
    assert any(b["id"] == batch_id for b in listing.json())

    detail = client.get(f"/delayed-cash/batches/{batch_id}", headers=_auth(admin_token))
    assert detail.status_code == 200
    assert detail.json()["id"] == batch_id

    assert client.get("/delayed-cash/batches/999999", headers=_auth(admin_token)).status_code == 404
    assert client.get("/delayed-cash/batches").status_code == 401


# ---------------------------------------------------------------------------
# Bulk publish
# ---------------------------------------------------------------------------


def test_publish_batch_issues_links_for_every_center_and_updates_status(client):
    _ensure_approved_rule("DCB-UPLOAD-PUBLISH")
    admin_token = _admin(client, "dcbu_admin_publish", "dcbu_admin_publish@example.com")

    rows = [
        ["903-A-C", "Publish Center A", "PB-A1", date(2026, 7, 1), datetime(2026, 7, 2, 9, 0), date(2026, 7, 2), 1, None, None],
        ["904-B-C", "Publish Center B", "PB-B1", date(2026, 7, 1), datetime(2026, 7, 3, 9, 0), date(2026, 7, 3), 2, None, None],
    ]
    upload_resp = _upload(client, admin_token, rows)
    batch_id = upload_resp.json()["batch"]["id"]

    publish = client.post(f"/delayed-cash/batches/{batch_id}/publish", headers=_auth(admin_token))
    assert publish.status_code == 200
    body = publish.json()
    assert body["batch_id"] == batch_id
    assert len(body["links"]) == 2
    codes = {link["centre_code"] for link in body["links"]}
    assert codes == {"903-A-C", "904-B-C"}
    for link in body["links"]:
        assert link["response_url"].endswith(f"/respond/delayed-cash/{link['response_token']}")

    batch_detail = client.get(f"/delayed-cash/batches/{batch_id}", headers=_auth(admin_token)).json()
    assert batch_detail["status"] == "published"

    # Every link actually resolves through the real public portal lookup.
    for link in body["links"]:
        public = client.get(f"/public/delayed-cash/cases/{link['response_token']}")
        assert public.status_code == 200
        assert public.json()["centre_code"] == link["centre_code"]


def test_publish_batch_again_reissues_fresh_tokens(client):
    _ensure_approved_rule("DCB-UPLOAD-REPUBLISH")
    admin_token = _admin(client, "dcbu_admin_republish", "dcbu_admin_republish@example.com")

    rows = [["905-C-C", "Republish Center", "PB-C1", date(2026, 7, 1), datetime(2026, 7, 2, 9, 0), date(2026, 7, 2), 1, None, None]]
    upload_resp = _upload(client, admin_token, rows)
    batch_id = upload_resp.json()["batch"]["id"]

    first = client.post(f"/delayed-cash/batches/{batch_id}/publish", headers=_auth(admin_token)).json()
    second = client.post(f"/delayed-cash/batches/{batch_id}/publish", headers=_auth(admin_token)).json()

    first_token = first["links"][0]["response_token"]
    second_token = second["links"][0]["response_token"]
    assert first_token != second_token

    assert client.get(f"/public/delayed-cash/cases/{first_token}").status_code == 404
    assert client.get(f"/public/delayed-cash/cases/{second_token}").status_code == 200


def test_publish_unknown_batch_404s(client):
    admin_token = _admin(client, "dcbu_admin_publish404", "dcbu_admin_publish404@example.com")
    resp = client.post("/delayed-cash/batches/999999/publish", headers=_auth(admin_token))
    assert resp.status_code == 404


# ---------------------------------------------------------------------------
# Read-only "view already-published links" (no re-minting)
# ---------------------------------------------------------------------------


def test_get_links_before_publish_is_empty(client):
    _ensure_approved_rule("DCB-UPLOAD-LINKS-EMPTY")
    admin_token = _admin(client, "dcbu_admin_links_empty", "dcbu_admin_links_empty@example.com")

    rows = [["906-D-C", "Links Center D", "PB-D1", date(2026, 7, 1), datetime(2026, 7, 2, 9, 0), date(2026, 7, 2), 1, None, None]]
    upload_resp = _upload(client, admin_token, rows)
    batch_id = upload_resp.json()["batch"]["id"]

    links = client.get(f"/delayed-cash/batches/{batch_id}/links", headers=_auth(admin_token))
    assert links.status_code == 200
    assert links.json()["links"] == []


def test_get_links_after_publish_matches_without_reissuing_tokens(client):
    _ensure_approved_rule("DCB-UPLOAD-LINKS-VIEW")
    admin_token = _admin(client, "dcbu_admin_links_view", "dcbu_admin_links_view@example.com")

    rows = [
        ["907-E-C", "Links Center E", "PB-E1", date(2026, 7, 1), datetime(2026, 7, 2, 9, 0), date(2026, 7, 2), 1, None, None],
        ["908-F-C", "Links Center F", "PB-F1", date(2026, 7, 1), datetime(2026, 7, 3, 9, 0), date(2026, 7, 3), 2, None, None],
    ]
    upload_resp = _upload(client, admin_token, rows)
    batch_id = upload_resp.json()["batch"]["id"]

    published = client.post(f"/delayed-cash/batches/{batch_id}/publish", headers=_auth(admin_token)).json()
    published_tokens = {link["centre_code"]: link["response_token"] for link in published["links"]}

    # Viewing links must NOT mint fresh tokens -- calling it twice in a row
    # returns the exact same tokens both times.
    first_view = client.get(f"/delayed-cash/batches/{batch_id}/links", headers=_auth(admin_token)).json()
    second_view = client.get(f"/delayed-cash/batches/{batch_id}/links", headers=_auth(admin_token)).json()

    assert first_view["batch_id"] == batch_id
    assert len(first_view["links"]) == 2
    for view in (first_view, second_view):
        view_tokens = {link["centre_code"]: link["response_token"] for link in view["links"]}
        assert view_tokens == published_tokens

    # And the already-published links still resolve through the public portal.
    for link in first_view["links"]:
        public = client.get(f"/public/delayed-cash/cases/{link['response_token']}")
        assert public.status_code == 200
        assert public.json()["centre_code"] == link["centre_code"]


def test_get_links_unknown_batch_404s(client):
    admin_token = _admin(client, "dcbu_admin_links404", "dcbu_admin_links404@example.com")
    resp = client.get("/delayed-cash/batches/999999/links", headers=_auth(admin_token))
    assert resp.status_code == 404
