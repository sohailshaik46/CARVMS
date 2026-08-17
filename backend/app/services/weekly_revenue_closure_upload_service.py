"""Parses the raw weekly Weekly Revenue Closure "pending list" workbook --
proven from a real reference file (`July-26-Week2-closure pending List till
12-6-2026.xlsx`, `Center wise` sheet: Zone, Cluster, Center Code, Center
Name, Date, Billed Sessions, Daily Report, Variance, Remark, Final
Remarks) -- into pending WeeklyRevenueBillIncident rows, awaiting a center
remark and a Vigilance verdict (mirrors the review-queue pattern already
built for Delayed Cash Billing: ingest now with considered=None, decide
later).

This is the piece that was missing when the formula was first proven
(docs/CARVMS_WEEKLY_REVENUE_CLOSURE_FORMULA_ANALYSIS.md S6.4) -- the two
Penalty *output* workbooks only ever contained each week's result, never
the raw daily source it was computed from. Verified against the real Week
2 pending file: excluding "Excess billed/Incorrect Daily report" rows (a
separate, non-penalized anomaly category -- present in the raw file but
never appearing in either Penalty output workbook), this format's own
per-center incident-type counts reproduce that same file's own `Center
Penalty` sheet exactly, 35/35 centers, zero mismatches.

Week 3's equivalent pending file does NOT reconcile against its own
`Center Penalty` sheet for 7 centers (all undercounts in the sheet, same
kind of pivot-staleness issue documented for Week 3's Penalty output) --
this parser reproduces the correct per-row counts from the raw sheet
regardless, per this project's standing rule to never quietly reproduce a
known-wrong figure.
"""

import io
from dataclasses import dataclass
from datetime import date, datetime
from typing import Optional

import openpyxl

from app.services import weekly_revenue_closure_service as calc_service

SHEET_NAME = "Center wise"

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "zone": ("zone",),
    "cluster": ("cluster",),
    "centre_code": ("centercode", "centrecode"),
    "centre_name": ("centername", "centrename"),
    "incident_date": ("date",),
    "billed_sessions": ("billedsessions",),
    "daily_report": ("dailyreport",),
    "variance": ("variance",),
    "raw_remark": ("remark",),
    "final_remarks": ("finalremarks", "misfinalremarks"),
}

REQUIRED_FIELDS = ("centre_code", "centre_name", "incident_date", "final_remarks")

# Proven from the real Week 2 pending file: this category exists in the
# raw daily source but never appears in either Penalty output workbook --
# it's a separate anomaly (billed sessions EXCEED the daily report, the
# opposite direction from a delayed-closure problem), out of scope for
# this penalty engine. Rows of this type are counted and reported, never
# silently dropped, but never turned into a penalty-eligible incident.
EXCESS_BILLED_TYPE = "excess_billed_or_incorrect_report"

FINAL_REMARK_TO_TYPE = {
    "billpending": "bill_pending",
    "dailyreportnotsent": "daily_report_not_sent",
    # "no billing /no daily report" normalizes to this after stripping
    # non-alphanumerics -- matched via substring below instead of an exact
    # key, since the real header text varies slightly between files.
}


def _normalize_header(text) -> str:
    return "".join(ch for ch in str(text).strip().lower() if ch.isalnum())


def _build_column_map(header_row) -> dict[str, int]:
    normalized = {_normalize_header(v): i for i, v in enumerate(header_row) if v is not None}
    column_map: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                column_map[field] = normalized[alias]
                break
    return column_map


def _classify_final_remark(value) -> Optional[str]:
    normalized = _normalize_header(value)
    if not normalized:
        return None
    if "excess" in normalized:
        return EXCESS_BILLED_TYPE
    if "nobilling" in normalized:
        return "no_billing_no_daily_report"
    return FINAL_REMARK_TO_TYPE.get(normalized)


def _coerce_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


@dataclass
class SkippedPendingRow:
    row_number: int
    reason: str


def parse_pending_workbook(
    raw_bytes: bytes,
) -> tuple[list[calc_service.RawBillIncidentInput], int, list[SkippedPendingRow]]:
    """Returns (penalty-eligible incidents, excess-billed row count seen but
    excluded, skipped rows). Never raises on a bad individual row -- only
    on a genuinely unusable header (mirrors delayed_cash_upload_service's
    contract)."""
    workbook = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    sheet_name = SHEET_NAME if SHEET_NAME in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], 0, []

    column_map = _build_column_map(rows[0])
    missing = [f for f in REQUIRED_FIELDS if f not in column_map]
    if missing:
        raise ValueError(
            f"'{sheet_name}' is missing required column(s): {', '.join(missing)}. Expected headers like "
            "Zone, Cluster, Center Code, Center Name, Date, Remark, Final Remarks."
        )

    def cell(row, field):
        idx = column_map.get(field)
        return row[idx] if idx is not None and idx < len(row) else None

    incidents: list[calc_service.RawBillIncidentInput] = []
    excess_billed_count = 0
    skipped: list[SkippedPendingRow] = []

    for row_number, row in enumerate(rows[1:], start=2):
        if row is None or all(v is None for v in row):
            continue
        try:
            centre_code = str(cell(row, "centre_code") or "").strip()
            centre_name = str(cell(row, "centre_name") or "").strip()
            if not centre_code or not centre_name:
                skipped.append(SkippedPendingRow(row_number, "Missing Center Code/Name"))
                continue

            incident_date = _coerce_date(cell(row, "incident_date"))
            if incident_date is None:
                skipped.append(SkippedPendingRow(row_number, "Unparseable Date"))
                continue

            incident_type = _classify_final_remark(cell(row, "final_remarks"))
            if incident_type is None:
                skipped.append(
                    SkippedPendingRow(row_number, f"Unrecognized Final Remarks: {cell(row, 'final_remarks')!r}")
                )
                continue
            if incident_type == EXCESS_BILLED_TYPE:
                excess_billed_count += 1
                continue

            incidents.append(
                calc_service.RawBillIncidentInput(
                    centre_code=centre_code,
                    centre_name=centre_name,
                    zone=cell(row, "zone"),
                    cluster=cell(row, "cluster"),
                    incident_date=incident_date,
                    mis_final_remark=incident_type,
                    billed_sessions=cell(row, "billed_sessions"),
                    daily_report=cell(row, "daily_report"),
                    variance=cell(row, "variance"),
                    raw_remark=cell(row, "raw_remark"),
                    # Deliberately no center_remarks/penalty_remarks yet --
                    # this is the pre-remark pending list. A center manager
                    # responds later (or doesn't); Vigilance decides the
                    # verdict from the review queue, same as Delayed Cash
                    # Billing.
                )
            )
        except Exception as exc:  # noqa: BLE001 -- one bad row must never abort the batch
            skipped.append(SkippedPendingRow(row_number, f"Unexpected error: {exc}"))
            continue

    return incidents, excess_billed_count, skipped
