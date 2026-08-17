"""Generates the multi-sheet Excel workbook proven in
docs/CARVMS_WEEKLY_REVENUE_CLOSURE_FORMULA_ANALYSIS.md -- reproduces the
real reference workbooks' three-sheet shape (an unfiltered incident-count
pivot, the two-section penalty breakdown with Cluster/Zonal rollups, and
the remark-received raw detail) from a closed batch's own data, never from
hand-maintained pivot tables that (per the same doc, S6.1) are shown to
contain real arithmetic errors in the source files this was reverse-
engineered from.
"""

import io
from collections import defaultdict
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.models.weekly_revenue_closure import (
    WeeklyRevenueBillIncident,
    WeeklyRevenueClosureBatch,
    WeeklyRevenueNoRemarkIncident,
    WeeklyRevenueRolePenalty,
)

_TYPE_LABELS = {
    "bill_pending": "Bill Pending",
    "daily_report_not_sent": "Daily Report not sent",
    "no_billing_no_daily_report": "No Billing /No Daily Report",
}

_BOLD = Font(bold=True)


def generate_penalty_workbook(db: Session, *, batch: WeeklyRevenueClosureBatch) -> bytes:
    bill_incidents = (
        db.query(WeeklyRevenueBillIncident).filter(WeeklyRevenueBillIncident.batch_id == batch.id).all()
    )
    no_remark_incidents = (
        db.query(WeeklyRevenueNoRemarkIncident).filter(WeeklyRevenueNoRemarkIncident.batch_id == batch.id).all()
    )
    role_penalties = (
        db.query(WeeklyRevenueRolePenalty).filter(WeeklyRevenueRolePenalty.batch_id == batch.id).all()
    )

    wb = openpyxl.Workbook()
    _write_sheet1(wb, bill_incidents, no_remark_incidents)
    _write_penalties_sheet(wb, batch, bill_incidents, no_remark_incidents, role_penalties)
    _write_data_sheet(wb, batch, bill_incidents)
    del wb["Sheet"]  # the default empty sheet Workbook() creates

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _write_sheet1(wb, bill_incidents, no_remark_incidents):
    """Unfiltered incident-count pivot, by center x MIS category --
    informational only, matches the proven `Sheet1` layout. Excludes any
    bill incident moved to "Remarks Not Received" (moved_to_no_remark) --
    it's already represented by its own no_remark_incidents row, so
    counting both would double-count the same real-world incident."""
    ws = wb.create_sheet("Sheet1")
    counts: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    names: dict[str, str] = {}
    for b in bill_incidents:
        if b.moved_to_no_remark:
            continue
        counts[b.centre_code][b.mis_final_remark] += 1
        names[b.centre_code] = b.centre_name
    for n in no_remark_incidents:
        counts[n.centre_code][n.incident_type] += n.incident_count
        names[n.centre_code] = n.centre_name

    ws.append(["Count of Center Code", None, "MIS Final Remarks"])
    header = ["Center Code", "Center Name", "Bill Pending", "Daily Report not sent", "No Billing /No Daily Report", "Grand Total"]
    ws.append(header)
    for cell in ws[2]:
        cell.font = _BOLD

    grand = defaultdict(int)
    for code in sorted(counts.keys()):
        row_counts = counts[code]
        bp = row_counts.get("bill_pending", 0)
        dr = row_counts.get("daily_report_not_sent", 0)
        nb = row_counts.get("no_billing_no_daily_report", 0)
        total = bp + dr + nb
        ws.append([code, names[code], bp or None, dr or None, nb or None, total])
        grand["bp"] += bp
        grand["dr"] += dr
        grand["nb"] += nb
        grand["total"] += total
    ws.append(["Grand Total", None, grand["bp"] or None, grand["dr"] or None, grand["nb"] or None, grand["total"]])
    ws["A" + str(ws.max_row)].font = _BOLD


def _cluster_rollup(role_penalties, role: str, section: str):
    return [rp for rp in role_penalties if rp.role == role and rp.section == section]


def _write_penalties_sheet(wb, batch, bill_incidents, no_remark_incidents, role_penalties):
    ws = wb.create_sheet("Penalties")
    row_cursor = 1

    def write_row(values, bold=False):
        nonlocal row_cursor
        ws.append(values)
        if bold:
            for cell in ws[row_cursor]:
                cell.font = _BOLD
        row_cursor += 1

    # ---- Section 1: Remarks received but not considered ----
    write_row([f"Remarks received but not considered - {batch.week_label}"], bold=True)
    write_row(["Center Code", "Center Name", "Center Manager", "NP ID", "Cluster", "Zone", "Zonal Manager", "Bill Pending", "Daily Report not sent", "Penalty"], bold=True)

    not_considered_by_center: dict[str, list] = defaultdict(list)
    for b in bill_incidents:
        if b.considered == "not_considered":
            not_considered_by_center[b.centre_code].append(b)

    for code in sorted(not_considered_by_center.keys()):
        rows = not_considered_by_center[code]
        first = rows[0]
        bp = sum(1 for r in rows if r.mis_final_remark == "bill_pending")
        dr = sum(1 for r in rows if r.mis_final_remark == "daily_report_not_sent")
        write_row([
            code, first.centre_name, first.center_manager, first.center_manager_npid, first.cluster,
            first.zone, first.zonal_manager, bp or None, dr or None, float(Decimal("0.0625")),
        ])
    write_row([])
    write_row(["Cluster Manager", "NP ID", "Count", "Penalty"], bold=True)
    for rp in sorted(_cluster_rollup(role_penalties, "cluster_manager", "not_considered"), key=lambda r: r.person_name):
        write_row([rp.person_name, rp.person_npid, rp.distinct_center_count, float(rp.penalty_amount)])
    write_row([])

    # ---- Section 2: Remarks Not Received ----
    write_row([f"Remarks Not Received - {batch.week_label}"], bold=True)
    write_row(["Center Code", "Center Name", "Center Manager", "NP ID", "Cluster", "Zone", "Zonal Manager", "Bill Pending", "Daily Report not sent", "No Billing /No Daily Report", "Penalty"], bold=True)

    no_remark_by_center: dict[str, list] = defaultdict(list)
    for n in no_remark_incidents:
        no_remark_by_center[n.centre_code].append(n)

    for code in sorted(no_remark_by_center.keys()):
        rows = no_remark_by_center[code]
        first = rows[0]
        by_type = {r.incident_type: r.incident_count for r in rows}
        write_row([
            code, first.centre_name, first.center_manager, first.center_manager_npid, first.cluster,
            first.zone, first.zonal_manager,
            by_type.get("bill_pending") or None, by_type.get("daily_report_not_sent") or None,
            by_type.get("no_billing_no_daily_report") or None, float(Decimal("0.0625")),
        ])
    write_row([])
    write_row(["Cluster Manager", "NP ID", "Count", "Penalty", None, "Zonal Manager", "NP ID", "Count", "Penalty"], bold=True)
    cluster_rows = sorted(_cluster_rollup(role_penalties, "cluster_manager", "no_remark"), key=lambda r: r.person_name)
    zonal_rows = sorted(_cluster_rollup(role_penalties, "zonal_manager", "no_remark"), key=lambda r: r.person_name)
    for i in range(max(len(cluster_rows), len(zonal_rows))):
        c = cluster_rows[i] if i < len(cluster_rows) else None
        z = zonal_rows[i] if i < len(zonal_rows) else None
        write_row([
            c.person_name if c else None, c.person_npid if c else None,
            c.distinct_center_count if c else None, float(c.penalty_amount) if c else None, None,
            z.person_name if z else None, z.person_npid if z else None,
            z.distinct_center_count if z else None, float(z.penalty_amount) if z else None,
        ])


def _write_data_sheet(wb, batch, bill_incidents):
    """Mirrors the proven 'Data' sheet -- remark-received incidents only.
    An incident that Vigilance has since moved to "Remarks Not Received"
    (via mark_no_remark_received) is excluded here even though its row is
    kept in the table as an audit trail -- it no longer counts as
    remark-received for export purposes (see model docstring)."""
    ws = wb.create_sheet("Data")
    ws.append([
        "S.No", "Zone", "Cluster", "Center Code", "Center Name", "Date", "Billed Sessions", "Daily Report",
        "Variance", "Remark", "MIS Final Remarks", "Center remarks", "Penalty Remarks", "Week",
    ])
    for cell in ws[1]:
        cell.font = _BOLD

    ordered = sorted(
        (b for b in bill_incidents if not b.moved_to_no_remark),
        key=lambda b: (b.incident_date, b.centre_code),
    )
    for i, b in enumerate(ordered, start=1):
        ws.append([
            i, b.zone, b.cluster, b.centre_code, b.centre_name, b.incident_date, b.billed_sessions,
            b.daily_report, b.variance, b.raw_remark, _TYPE_LABELS.get(b.mis_final_remark, b.mis_final_remark),
            b.center_remarks, b.penalty_remarks, batch.week_label,
        ])
