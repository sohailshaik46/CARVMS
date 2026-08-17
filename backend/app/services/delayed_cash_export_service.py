"""Generates the Delayed Cash Billing penalty workbook -- a "Data" sheet
(every bill, raw + derived columns) and a "Penalty" sheet (one row per
center, the aggregate penalty table) -- regenerated from a batch's own
computed rows, never from a hand-maintained pivot. Mirrors the pattern in
weekly_revenue_closure_export_service.py; kept as a separate module since
the two engines' formulas and role hierarchies must never be merged.
"""

import io

import openpyxl
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.models.delayed_cash_billing import DelayedCashBill, DelayedCashCenterPenalty, DelayedCashUploadBatch

_BOLD = Font(bold=True)

_DATA_HEADERS = [
    "Center Code", "Center Name", "Sales Bill", "Bill Date", "Bill Created Time", "Created Date",
    "Source Day Difference", "Calculated Day Difference", "Difference Check", "Data Quality Status",
    "Calculated Penalty", "Center Remarks", "Penalty Remarks", "Considered",
]

_PENALTY_HEADERS = [
    "Center Code", "Center Name", "Total Bills", "Calculated Penalty", "Validated Penalty",
    "Monthly Cap Amount", "Final Penalty", "Penalty Status",
]


def generate_penalty_workbook(db: Session, *, batch: DelayedCashUploadBatch) -> bytes:
    bills = db.query(DelayedCashBill).filter(DelayedCashBill.batch_id == batch.id).order_by(
        DelayedCashBill.centre_code, DelayedCashBill.bill_date
    ).all()
    center_penalties = (
        db.query(DelayedCashCenterPenalty)
        .filter(DelayedCashCenterPenalty.batch_id == batch.id)
        .order_by(DelayedCashCenterPenalty.centre_name)
        .all()
    )

    wb = openpyxl.Workbook()
    _write_data_sheet(wb, bills)
    _write_penalty_sheet(wb, center_penalties)
    del wb["Sheet"]  # the default empty sheet Workbook() creates

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _write_data_sheet(wb, bills: list[DelayedCashBill]):
    ws = wb.create_sheet("Data")
    ws.append(_DATA_HEADERS)
    for cell in ws[1]:
        cell.font = _BOLD

    for b in bills:
        ws.append([
            b.centre_code, b.centre_name, b.sales_bill, b.bill_date, b.bill_created_time, b.created_date,
            b.source_day_difference, b.calculated_day_difference, b.difference_check, b.data_quality_status,
            float(b.calculated_penalty), b.center_remarks, b.penalty_remarks, b.considered,
        ])


def _write_penalty_sheet(wb, center_penalties: list[DelayedCashCenterPenalty]):
    """One row per center -- the actual "penalty table" this format is
    proven against (see formula analysis doc)."""
    ws = wb.create_sheet("Penalty")
    ws.append(_PENALTY_HEADERS)
    for cell in ws[1]:
        cell.font = _BOLD

    total_calculated = 0.0
    total_final = 0.0
    for cp in center_penalties:
        calculated = float(cp.calculated_penalty)
        final = float(cp.final_penalty) if cp.final_penalty is not None else None
        total_calculated += calculated
        if final is not None:
            total_final += final
        ws.append([
            cp.centre_code, cp.centre_name, cp.total_bills, calculated,
            float(cp.validated_penalty) if cp.validated_penalty is not None else None,
            float(cp.monthly_cap_amount) if cp.monthly_cap_amount is not None else None,
            final, cp.penalty_status,
        ])

    ws.append(["Grand Total", None, None, total_calculated, None, None, total_final or None, None])
    for cell in ws[ws.max_row]:
        cell.font = _BOLD
