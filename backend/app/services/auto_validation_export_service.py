"""Center/zone/cluster-wise Excel export of auto-validated response remarks,
for both engines -- 4 sheets: a flat "Raw Data" row per response (with a
repeat-instance count so a recurring excuse from the same center stands
out), plus "By Center" / "By Zone" / "By Cluster" bucket-count rollups.

DCB responses carry only centre_code natively -- zone/cluster for those come
from the Org Master (same asymmetry documented in
metrics.py's _cluster_zone_breakdown; a DCB center not yet linked there
shows as "Unknown", same convention). WRC responses resolve zone/cluster
from their own batch's incident rows, which already carry those columns
verbatim from the uploaded sheet.
"""

import io
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.models.weekly_revenue_closure import WeeklyRevenueBillIncident
from app.services import auto_validation_service
from app.services import org_service

_BOLD = Font(bold=True)
_UNKNOWN = "Unknown"


@dataclass
class AutoValidationRow:
    engine: str  # "DCB" | "WRC"
    response_id: int
    zone: str
    cluster: str
    centre_code: str
    centre_name: str
    category: Optional[str]
    matched_keyword: Optional[str]
    auto_bucket: Optional[str]
    effective_bucket: Optional[str]
    decision_label: Optional[str]
    reason: Optional[str]
    admin_overridden: bool
    submitted_at: datetime


def _resolve_dcb_zone_cluster(db: Session, centre_code: str) -> tuple[str, str]:
    node = org_service.get_node_by_external_code(db, centre_code)
    if node is None:
        return _UNKNOWN, _UNKNOWN
    zone_node = org_service.find_ancestor_by_dimension_key(db, node, "zone")
    cluster_node = org_service.find_ancestor_by_dimension_key(db, node, "cluster")
    return (zone_node.name if zone_node else _UNKNOWN, cluster_node.name if cluster_node else _UNKNOWN)


def build_dcb_rows(db: Session) -> list[AutoValidationRow]:
    rows = []
    for r in auto_validation_service.list_dcb_responses(db):
        cp = r.center_penalty
        zone, cluster = _resolve_dcb_zone_cluster(db, cp.centre_code)
        rows.append(
            AutoValidationRow(
                engine="DCB",
                response_id=r.id,
                zone=zone,
                cluster=cluster,
                centre_code=cp.centre_code,
                centre_name=cp.centre_name,
                category=r.auto_category,
                matched_keyword=r.auto_matched_keyword,
                auto_bucket=r.auto_bucket,
                effective_bucket=auto_validation_service.effective_bucket(r),
                decision_label=r.auto_decision_label,
                reason=r.auto_reason,
                admin_overridden=r.admin_override_bucket is not None,
                submitted_at=r.submitted_at,
            )
        )
    return rows


def build_wrc_rows(db: Session) -> list[AutoValidationRow]:
    rows = []
    for r in auto_validation_service.list_wrc_responses(db):
        case = r.case
        incident = (
            db.query(WeeklyRevenueBillIncident)
            .filter(
                WeeklyRevenueBillIncident.batch_id == case.batch_id,
                WeeklyRevenueBillIncident.centre_code == case.centre_code,
            )
            .first()
        )
        rows.append(
            AutoValidationRow(
                engine="WRC",
                response_id=r.id,
                zone=(incident.zone if incident and incident.zone else _UNKNOWN),
                cluster=(incident.cluster if incident and incident.cluster else _UNKNOWN),
                centre_code=case.centre_code,
                centre_name=case.centre_name,
                category=r.auto_category,
                matched_keyword=r.auto_matched_keyword,
                auto_bucket=r.auto_bucket,
                effective_bucket=auto_validation_service.effective_bucket(r),
                decision_label=r.auto_decision_label,
                reason=r.auto_reason,
                admin_overridden=r.admin_override_bucket is not None,
                submitted_at=r.submitted_at,
            )
        )
    return rows


def _repeat_counts(rows: list[AutoValidationRow]) -> dict:
    """How many times this exact (engine, center, category) combination
    recurs across every response pulled -- the "repeated instances" signal
    the user asked for, e.g. a center that has claimed "Staff on leave"
    five separate times stands out immediately instead of blending into
    a flat list."""
    return Counter((r.engine, r.centre_code, r.category) for r in rows if r.category)


_RAW_HEADERS = [
    "Engine", "Zone", "Cluster", "Centre Code", "Centre Name", "Category", "Matched Keyword",
    "Auto Decision", "Effective Decision", "Admin Overridden?", "Reason (if not considered)",
    "Repeat Instances (same center + category)", "Submitted At",
]


def _write_raw_sheet(wb, rows: list[AutoValidationRow], repeat_counts: dict):
    ws = wb.create_sheet("Raw Data")
    ws.append(_RAW_HEADERS)
    for cell in ws[1]:
        cell.font = _BOLD
    for r in rows:
        repeat = repeat_counts.get((r.engine, r.centre_code, r.category), 1) if r.category else 1
        ws.append(
            [
                r.engine, r.zone, r.cluster, r.centre_code, r.centre_name,
                r.category or "(no rule matched)", r.matched_keyword,
                r.auto_bucket, r.effective_bucket, "Yes" if r.admin_overridden else "No",
                r.reason, repeat, r.submitted_at,
            ]
        )


def _write_rollup_sheet(wb, sheet_name: str, rows: list[AutoValidationRow], key_fn):
    """One row per distinct (engine, key) -- total responses + a count per
    effective bucket, so "which centers/zones/clusters have the most
    not_considered or manual_check remarks" is readable at a glance without
    a pivot table."""
    ws = wb.create_sheet(sheet_name)
    headers = ["Engine", sheet_name.replace("By ", ""), "Total", "Considered", "Not Considered", "Manual Check"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _BOLD

    groups: dict = {}
    for r in rows:
        key = (r.engine, key_fn(r))
        bucket_counts = groups.setdefault(key, {"considered": 0, "not_considered": 0, "manual_check": 0, "total": 0})
        bucket_counts["total"] += 1
        if r.effective_bucket in bucket_counts:
            bucket_counts[r.effective_bucket] += 1

    for (engine, key), counts in sorted(groups.items(), key=lambda kv: (kv[0][0], -kv[1]["total"])):
        ws.append(
            [engine, key, counts["total"], counts["considered"], counts["not_considered"], counts["manual_check"]]
        )


def render_xlsx(dcb_rows: list[AutoValidationRow], wrc_rows: list[AutoValidationRow]) -> bytes:
    all_rows = dcb_rows + wrc_rows
    repeat_counts = _repeat_counts(all_rows)

    wb = openpyxl.Workbook()
    _write_raw_sheet(wb, all_rows, repeat_counts)
    _write_rollup_sheet(wb, "By Center", all_rows, lambda r: f"{r.centre_code} - {r.centre_name}")
    _write_rollup_sheet(wb, "By Zone", all_rows, lambda r: r.zone)
    _write_rollup_sheet(wb, "By Cluster", all_rows, lambda r: r.cluster)
    del wb["Sheet"]

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
