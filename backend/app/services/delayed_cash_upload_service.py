"""Parses a raw weekly Delayed Cash Billing bills workbook -- the `Bills
Data` sheet layout proven in
docs/CARVMS_DELAYED_CASH_PENALTY_FORMULA_ANALYSIS.md (CENTREID, CENTRENAME,
SALESBILL, BILLDATE, bill_created_time, created_date, day_difference,
Center Remarks, Penalty Remarks) -- into RawBillInput rows, then
orchestrates the existing, already-proven calculator pipeline
(create_upload_batch -> ingest_bills -> compute_center_penalties) against
them.

Never invents a row: one that's missing a mandatory field, has an
unparseable date, a non-numeric delay, or repeats a SALESBILL already seen
earlier in the same file is skipped and reported by row number -- never
silently dropped, and never allowed to abort the rest of the batch (mirrors
the per-row isolation pattern in org_sheet_sync_service.py)."""

import io
from dataclasses import dataclass
from datetime import date, datetime, timezone
from typing import Optional

import openpyxl
from sqlalchemy.orm import Session

from app.models.delayed_cash_billing import (
    DelayedCashCenterPenalty,
    DelayedCashPenaltyRule,
    DelayedCashUploadBatch,
)
from app.models.user import User
from app.services import delayed_cash_penalty_service as calc_service
from app.services import delayed_cash_response_service as response_service

SHEET_NAME = "Bills Data"

# Canonical field <- accepted header aliases. Matching is case/space/
# punctuation-insensitive (see _normalize_header) so "day_difference",
# "Day Difference" and "DAYDIFFERENCE" all resolve the same way.
HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "centre_code": ("centreid", "centrecode", "centercode"),
    "centre_name": ("centrename", "centername"),
    "sales_bill": ("salesbill", "billno", "billnumber"),
    "bill_date": ("billdate",),
    "bill_created_time": ("billcreatedtime", "createdtime"),
    "created_date": ("createddate",),
    "source_day_difference": ("daydifference", "delay"),
    "center_remarks": ("centerremarks", "centreremarks"),
    "penalty_remarks": ("penaltyremarks",),
}

REQUIRED_FIELDS = (
    "centre_code",
    "centre_name",
    "sales_bill",
    "bill_date",
    "created_date",
    "source_day_difference",
)


def _normalize_header(text) -> str:
    return "".join(ch for ch in str(text).strip().lower() if ch.isalnum())


def _build_column_map(header_row) -> dict[str, int]:
    normalized = {_normalize_header(v): i for i, v in enumerate(header_row) if v is not None}
    column_map: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                column_map[field] = normalized[alias]
                break
    return column_map


@dataclass
class SkippedBillRow:
    row_number: int  # 1-based, matching the spreadsheet's own row numbers (header = row 1)
    reason: str


def _coerce_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _coerce_datetime(value) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    parsed = _coerce_date(value)
    if parsed is None:
        return None
    return datetime.combine(parsed, datetime.min.time(), tzinfo=timezone.utc)


def _find_data_sheet(workbook) -> tuple[str, list, dict[str, int]]:
    """Finds whichever sheet actually contains every required column --
    the real source files aren't consistent about which sheet the data
    lands on (it's been "Bills Data", "Sheet1", and others), so this
    checks the preferred name first, then every other sheet in the
    workbook, rather than assuming a fixed position. Returns the sheet's
    name, its rows, and its column map. Raises ValueError only if NO
    sheet in the workbook has all the required columns."""
    ordered_names = [SHEET_NAME] if SHEET_NAME in workbook.sheetnames else []
    ordered_names += [n for n in workbook.sheetnames if n not in ordered_names]

    checked: list[tuple[str, list[str]]] = []
    for name in ordered_names:
        sheet = workbook[name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows or rows[0] is None:
            continue
        column_map = _build_column_map(rows[0])
        missing = [f for f in REQUIRED_FIELDS if f not in column_map]
        if not missing:
            return name, rows, column_map
        checked.append((name, missing))

    detail = "; ".join(f"'{name}' missing {', '.join(m)}" for name, m in checked) or "the workbook has no sheets"
    raise ValueError(
        f"No sheet in this workbook has every required column ({detail}). Expected headers like "
        "CENTREID, CENTRENAME, SALESBILL, BILLDATE, created_date, day_difference -- on any sheet."
    )


def parse_bills_workbook(raw_bytes: bytes) -> tuple[list[calc_service.RawBillInput], list[SkippedBillRow]]:
    """Searches every sheet in the workbook for the one containing all the
    required columns (see _find_data_sheet) and returns (parsed rows,
    skipped rows). Raises ValueError only when no sheet has a usable
    header row at all -- a missing/bad individual data row never raises,
    it's skipped."""
    workbook = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    sheet_name, rows, column_map = _find_data_sheet(workbook)
    if len(rows) <= 1:
        return [], []

    parsed: list[calc_service.RawBillInput] = []
    skipped: list[SkippedBillRow] = []
    seen_sales_bills: set[str] = set()

    def cell(row, field):
        idx = column_map.get(field)
        return row[idx] if idx is not None and idx < len(row) else None

    for row_number, row in enumerate(rows[1:], start=2):
        if row is None or all(v is None for v in row):
            continue  # a genuinely blank spreadsheet row, not a data error

        try:
            centre_code = str(cell(row, "centre_code") or "").strip()
            centre_name = str(cell(row, "centre_name") or "").strip()
            sales_bill = str(cell(row, "sales_bill") or "").strip()
            if not centre_code or not centre_name or not sales_bill:
                skipped.append(SkippedBillRow(row_number, "Missing Center Code/Name or Sales Bill number"))
                continue
            if sales_bill in seen_sales_bills:
                skipped.append(SkippedBillRow(row_number, f"Duplicate Sales Bill '{sales_bill}' within this file"))
                continue

            bill_date = _coerce_date(cell(row, "bill_date"))
            created_date = _coerce_date(cell(row, "created_date"))
            if bill_date is None or created_date is None:
                skipped.append(SkippedBillRow(row_number, "Unparseable BILLDATE or created_date"))
                continue

            bill_created_time = _coerce_datetime(cell(row, "bill_created_time")) or _coerce_datetime(created_date)

            raw_dd = cell(row, "source_day_difference")
            try:
                source_day_difference = int(raw_dd)
            except (TypeError, ValueError):
                skipped.append(SkippedBillRow(row_number, f"Non-numeric day_difference: {raw_dd!r}"))
                continue

            seen_sales_bills.add(sales_bill)
            center_remarks = cell(row, "center_remarks")
            penalty_remarks = cell(row, "penalty_remarks")
            parsed.append(
                calc_service.RawBillInput(
                    centre_code=centre_code,
                    centre_name=centre_name,
                    sales_bill=sales_bill,
                    bill_date=bill_date,
                    bill_created_time=bill_created_time,
                    created_date=created_date,
                    source_day_difference=source_day_difference,
                    center_remarks=str(center_remarks).strip() if center_remarks else None,
                    penalty_remarks=str(penalty_remarks).strip() if penalty_remarks else None,
                )
            )
        except Exception as exc:  # noqa: BLE001 -- one malformed row must never abort the batch
            skipped.append(SkippedBillRow(row_number, f"Unexpected error: {exc}"))
            continue

    return parsed, skipped


def upload_batch(
    db: Session,
    *,
    raw_bytes: bytes,
    source_filename: str,
    period_start: date,
    period_end: date,
    rule: DelayedCashPenaltyRule,
    uploaded_by: User,
) -> tuple[DelayedCashUploadBatch, list[DelayedCashCenterPenalty], list[SkippedBillRow]]:
    """Parses the workbook, then runs the proven calculator pipeline against
    every row that parsed cleanly. A batch is always created even if every
    row is skipped, so the skip report has somewhere to attach to."""
    raw_bills, skipped = parse_bills_workbook(raw_bytes)

    batch = calc_service.create_upload_batch(
        db,
        period_start=period_start,
        period_end=period_end,
        source_filename=source_filename,
        rule=rule,
        uploaded_by=uploaded_by,
    )
    calc_service.ingest_bills(db, batch=batch, rule=rule, raw_bills=raw_bills)
    center_penalties = calc_service.compute_center_penalties(db, batch=batch, rule=rule)
    return batch, center_penalties, skipped


def publish_batch(db: Session, *, batch: DelayedCashUploadBatch) -> list[DelayedCashCenterPenalty]:
    """Bulk-issues a fresh public response-portal token for every center
    penalty in this batch and marks the batch 'published'. Safe to call
    again -- always mints new tokens (invalidating any previous links),
    same contract as the single-case endpoint."""
    for cp in batch.center_penalties:
        response_service.generate_response_link_token(db, center_penalty=cp)
    batch.status = "published"
    db.commit()
    db.refresh(batch)
    return list(batch.center_penalties)


def get_published_links(db: Session, *, batch: DelayedCashUploadBatch) -> list[DelayedCashCenterPenalty]:
    """Read-only -- returns whichever centers already have a response
    token, WITHOUT minting or invalidating anything (unlike publish_batch
    above). Backs the Batches table's quick "View links" action, so
    copying a link doesn't require re-publishing (and so doesn't
    invalidate every other center's already-shared link) just to look."""
    return [cp for cp in batch.center_penalties if cp.response_token]
