import csv
import io

import openpyxl
import httpx
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile
from sqlalchemy.orm import Session

from app.auth import roles
from app.auth.dependencies import get_current_user, require_role
from app.config.settings import settings
from app.database.database import get_db
from app.models.user import User
from app.schemas.org import (
    CenterDetailOut,
    CenterDirectoryEntry,
    ContactChangeRequestOut,
    DataConflictOut,
    DirectoryReportOut,
    EmailSyncReportOut,
    OrgDimensionCreate,
    OrgDimensionOut,
    OrgNodeCreate,
    OrgNodeOut,
    OrgNodeUpdate,
    OrgNodeWithPath,
    SkippedRowOut,
    SyncReportOut,
)
from app.services import org_contact_change_service, org_service, org_sheet_sync_service
from app.services.org_service import (
    DimensionNotFoundError,
    DuplicateDimensionKeyError,
    DuplicateSiblingNameError,
    ParentNodeNotFoundError,
)

router = APIRouter(prefix="/org", tags=["Organization Hierarchy"])


def _report_to_schema(report) -> SyncReportOut:
    return SyncReportOut(
        total_rows=report.total_rows,
        half_countries_created=report.half_countries_created,
        zones_created=report.zones_created,
        clusters_created=report.clusters_created,
        centers_created=report.centers_created,
        centers_updated=report.centers_updated,
        skipped=[SkippedRowOut(row_number=s.row_number, reason=s.reason) for s in report.skipped],
        conflicts=[DataConflictOut(description=c.description) for c in report.conflicts],
    )


@router.get("/dimensions", response_model=list[OrgDimensionOut])
def list_dimensions(
    db: Session = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    return org_service.list_dimensions(db)


@router.post("/dimensions", response_model=OrgDimensionOut, status_code=201)
def create_dimension(
    payload: OrgDimensionCreate,
    db: Session = Depends(get_db),
    _admin: User = Depends(require_role(roles.ADMIN)),
):
    try:
        return org_service.create_dimension(
            db, key=payload.key, label=payload.label, sort_order=payload.sort_order
        )
    except DuplicateDimensionKeyError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.get("/nodes", response_model=list[OrgNodeOut])
def list_nodes(
    dimension_key: str | None = None,
    parent_id: int | None = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=5000),
    db: Session = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    return org_service.list_nodes(
        db, dimension_key=dimension_key, parent_id=parent_id, skip=skip, limit=limit
    )


@router.get("/centers-directory", response_model=list[CenterDirectoryEntry])
def get_centers_directory(
    db: Session = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    """Every active center as a flat {code, name} list, uncapped -- what the
    global center-search combobox is populated from everywhere in the app.
    Deliberately separate from GET /nodes (which is paginated/capped at 500
    and returns full node records) since a search box needs the whole
    center universe up front, not one page of the general hierarchy tree.
    Mirrors the public portal's identical directory (see
    org_sheet_sync_service.list_active_center_directory)."""
    entries = org_sheet_sync_service.list_active_center_directory(db)
    return [CenterDirectoryEntry(code=e["code"], name=e["name"]) for e in entries]


@router.get("/nodes/{node_id}", response_model=OrgNodeWithPath)
def get_node(
    node_id: int,
    db: Session = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    node = org_service.get_node(db, node_id)
    if node is None:
        raise HTTPException(status_code=404, detail="Node not found")
    return OrgNodeWithPath(
        id=node.id,
        dimension_id=node.dimension_id,
        parent_id=node.parent_id,
        name=node.name,
        external_code=node.external_code,
        is_active=node.is_active,
        manager_name=node.manager_name,
        manager_email=node.manager_email,
        manager_phone=node.manager_phone,
        manager_npid=node.manager_npid,
        path=org_service.get_node_path(db, node),
    )


@router.get("/centers/{center_code}/detail", response_model=CenterDetailOut)
def get_center_detail(
    center_code: str,
    db: Session = Depends(get_db),
    _user: User = Depends(get_current_user),
):
    """One flattened view of everything about a center -- its own manager
    (name/NPID/email/phone), its Cluster Manager, its Zonal Manager and
    zone name, and its Half Country Head, resolved by walking up the Org
    Master from the center node. 404s if the code isn't in the Org Master
    at all (see org_sheet_sync_service for how centers get placed there)."""
    detail = org_service.get_center_detail(db, center_code)
    if detail is None:
        raise HTTPException(status_code=404, detail=f"No center with code '{center_code}' in the Org Master")
    return detail


@router.post("/nodes", response_model=OrgNodeOut, status_code=201)
def create_node(
    payload: OrgNodeCreate,
    db: Session = Depends(get_db),
    _admin: User = Depends(require_role(roles.ADMIN)),
):
    try:
        return org_service.create_node(
            db,
            dimension_id=payload.dimension_id,
            parent_id=payload.parent_id,
            name=payload.name,
            external_code=payload.external_code,
        )
    except DuplicateSiblingNameError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except (DimensionNotFoundError, ParentNodeNotFoundError) as exc:
        raise HTTPException(status_code=404, detail=str(exc))


@router.patch("/nodes/{node_id}", response_model=OrgNodeOut)
def update_node(
    node_id: int,
    payload: OrgNodeUpdate,
    db: Session = Depends(get_db),
    _admin: User = Depends(require_role(roles.ADMIN)),
):
    node = org_service.get_node(db, node_id)
    if node is None:
        raise HTTPException(status_code=404, detail="Node not found")
    try:
        return org_service.update_node(
            db,
            node=node,
            name=payload.name,
            external_code=payload.external_code,
            is_active=payload.is_active,
            manager_name=payload.manager_name,
            manager_email=payload.manager_email,
            manager_phone=payload.manager_phone,
            manager_npid=payload.manager_npid,
        )
    except DuplicateSiblingNameError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/sync/centers-master", response_model=SyncReportOut)
async def sync_centers_master_upload(
    file: UploadFile,
    db: Session = Depends(get_db),
    _admin: User = Depends(require_role(roles.ADMIN)),
):
    """Manual upload path -- works today regardless of the Google Sheet's
    sharing settings. Export the sheet as CSV (File -> Download -> CSV) and
    upload it here. Reconciles the Org Master; never deletes a node that's
    missing from the file (see org_sheet_sync_service for why)."""
    raw = await file.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")

    delimiter = "\t" if (file.filename or "").lower().endswith((".tsv", ".txt")) else ","
    try:
        report = org_sheet_sync_service.sync_centers_master(db, text, delimiter=delimiter)
    except DimensionNotFoundError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return _report_to_schema(report)


@router.post("/sync/centers-master/from-sheet", response_model=SyncReportOut)
def sync_centers_master_from_sheet(
    db: Session = Depends(get_db),
    _admin: User = Depends(require_role(roles.ADMIN)),
):
    """Fetches the configured Google Sheet's CSV export directly and syncs
    it. Requires CENTERS_MASTER_SHEET_CSV_URL to be set AND the sheet to be
    shared as "Anyone with the link -- Viewer" -- a private sheet returns
    401 from Google, surfaced here as a clear error rather than a silent
    failure."""
    if not settings.CENTERS_MASTER_SHEET_CSV_URL:
        raise HTTPException(
            status_code=400,
            detail="CENTERS_MASTER_SHEET_CSV_URL is not configured -- set it in backend/.env, "
            "or use POST /org/sync/centers-master to upload a CSV export manually.",
        )
    try:
        response = httpx.get(settings.CENTERS_MASTER_SHEET_CSV_URL, timeout=30.0, follow_redirects=True)
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=502, detail=f"Could not reach the sheet: {exc}")

    if response.status_code == 401 or "accounts.google.com" in str(response.url):
        raise HTTPException(
            status_code=502,
            detail="Google returned 401/a login redirect -- the sheet is not shared as "
            "'Anyone with the link – Viewer' yet. Share it that way, or use "
            "POST /org/sync/centers-master to upload a CSV export manually instead.",
        )
    if response.status_code != 200:
        raise HTTPException(status_code=502, detail=f"Sheet fetch returned HTTP {response.status_code}")

    try:
        report = org_sheet_sync_service.sync_centers_master(db, response.text, delimiter=",")
    except DimensionNotFoundError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return _report_to_schema(report)


@router.post("/sync/center-directory", response_model=DirectoryReportOut)
async def sync_center_directory_upload(
    file: UploadFile,
    db: Session = Depends(get_db),
    _admin: User = Depends(require_role(roles.ADMIN)),
):
    """Upload the simpler 'All Centers' directory (Center ID / Center Name /
    Phone / CM Name / CM Email / Center Status / Address) -- keeps every
    center's basic identity current independent of whether the bigger
    Centers Master hierarchy sync has run. Powers the Center Code/Name
    dropdowns on the public response portal (see /public/delayed-cash/
    centers-directory)."""
    raw = await file.read()
    filename = (file.filename or "").lower()
    try:
        if filename.endswith((".xlsx", ".xlsm")):
            workbook = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
            sheet = workbook[workbook.sheetnames[0]]
            rows = list(sheet.iter_rows(values_only=True))
        else:
            text = raw.decode("utf-8-sig")
            rows = list(csv.reader(io.StringIO(text)))
    except Exception as exc:  # noqa: BLE001 -- surfaced as a clean 400, not a 500
        raise HTTPException(status_code=400, detail=f"Could not parse file: {exc}")

    try:
        report = org_sheet_sync_service.sync_center_directory(db, rows)
    except DimensionNotFoundError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    return DirectoryReportOut(
        total_rows=report.total_rows,
        centers_created=report.centers_created,
        centers_updated=report.centers_updated,
        skipped=[SkippedRowOut(row_number=s.row_number, reason=s.reason) for s in report.skipped],
    )


@router.post("/sync/center-emails", response_model=EmailSyncReportOut)
async def sync_center_emails_upload(
    file: UploadFile,
    db: Session = Depends(get_db),
    _admin: User = Depends(require_role(roles.ADMIN)),
):
    """Upload a center-email directory -- just Center Code + Email ID (plus
    whatever else the sheet has; matched by column name, not position, so
    extra/reordered columns like S.No or Region are fine). Updates
    OrgNode.manager_email for whichever center each row's code already
    resolves to in the Org Master -- never creates a new center (this
    sheet has no Zone/Cluster info to place one correctly). Powers the
    "notify center" email feature's To-address for every center."""
    raw = await file.read()
    filename = (file.filename or "").lower()
    try:
        if filename.endswith((".xlsx", ".xlsm")):
            workbook = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
            sheet = workbook[workbook.sheetnames[0]]
            rows = list(sheet.iter_rows(values_only=True))
        else:
            text = raw.decode("utf-8-sig")
            rows = list(csv.reader(io.StringIO(text)))
    except Exception as exc:  # noqa: BLE001 -- surfaced as a clean 400, not a 500
        raise HTTPException(status_code=400, detail=f"Could not parse file: {exc}")

    try:
        report = org_sheet_sync_service.sync_center_emails(db, rows)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    return EmailSyncReportOut(
        total_rows=report.total_rows,
        updated=report.updated,
        unchanged=report.unchanged,
        skipped=[SkippedRowOut(row_number=s.row_number, reason=s.reason) for s in report.skipped],
    )


# ---------------------------------------------------------------------------
# Center-manager contact-change notifications -- a public response-portal
# submission's name/NPID/email never writes to OrgNode directly (see
# org_contact_change_service); an Admin reviews and approves/rejects here.
# ---------------------------------------------------------------------------


@router.get("/contact-change-requests", response_model=list[ContactChangeRequestOut])
def list_contact_change_requests(
    status: str | None = Query(default=None, pattern="^(pending|approved|rejected)$"),
    db: Session = Depends(get_db),
    _admin: User = Depends(require_role(roles.ADMIN)),
):
    return org_contact_change_service.list_requests(db, status=status)


@router.post("/contact-change-requests/{request_id}/approve", response_model=ContactChangeRequestOut)
def approve_contact_change_request(
    request_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_role(roles.ADMIN)),
):
    try:
        request = org_contact_change_service.get_request_or_raise(db, request_id)
        return org_contact_change_service.approve_request(db, request=request, approver=admin)
    except org_contact_change_service.ContactChangeRequestNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except org_contact_change_service.NoMatchingOrgNodeError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except org_contact_change_service.AlreadyReviewedError as exc:
        raise HTTPException(status_code=409, detail=str(exc))


@router.post("/contact-change-requests/{request_id}/reject", response_model=ContactChangeRequestOut)
def reject_contact_change_request(
    request_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_role(roles.ADMIN)),
):
    try:
        request = org_contact_change_service.get_request_or_raise(db, request_id)
        return org_contact_change_service.reject_request(db, request=request, approver=admin)
    except org_contact_change_service.ContactChangeRequestNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except org_contact_change_service.AlreadyReviewedError as exc:
        raise HTTPException(status_code=409, detail=str(exc))
